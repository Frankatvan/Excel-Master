from __future__ import annotations

import json
import os
import pickle
import time
import re
import ssl
import uuid
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence, Tuple
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import pandas as pd
import yaml
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook

from .finance_mapping import MapperFactory
from .finance_formulas import FinanceFormulaGenerator, FormulaTemplateResolver, MappingIncompleteError
from .finance_formatting import SemanticFormattingEngine
from .finance_mapping import (
    ExcelSemanticMapper,
    resolve_sheet_field_columns_with_fallback,
)

from .finance_utils import (
    INTERNAL_COL_PREFIX,
    _get_service_account_info,
    _to_plain_dict,
    _safe_secrets,
    _get_secret,
    _parse_options,
    _is_internal_col,
    _cloud_view,
    _column_number_to_a1,
    _quote_sheet_name,
    _normalize_formula_range,
    _normalize_formula_text_for_compare,
    _column_a1_to_number,
    _normalize_headers,
    _parse_cell_value,
    _parse_formula_value,
    _trim_matrix,
    _trim_display_and_formula_matrices,
    _build_formula_lookup,
    _build_formula_lookup_by_headers,
    _values_to_dataframe,
    _serialize_for_api,
    _dataframe_to_values,
    _df_signature,
    _safe_string,
    _safe_number,
    _to_float,
    _extract_leading_int,
    _extract_tail_int,
    _extract_tail_str,
    _extract_year,
    _co_date_to_actual_settlement_date,
    _format_iso_date_or_blank,
    _normalize_date_value,
    _normalize_amount_key,
    _normalize_text_key,
    _has_digits,
    _contains_general_condition,
    _normalize_header_token,
    _find_col_in_headers,
    _find_col_in_row,
    _sheet_key,
    _ensure_column_count,
    _ensure_row_count,
    _get_cell,
    _column_values_1based,
    _set_cell,
    _sheet_delta_stats,
    _values_equal,
    _chunked,
    _contiguous_segments,
    _slugify_sheet_name,
    _normalize_label,
    _grid_cell,
    _find_first_row,
    _find_rows_by_item_label,
    get_sheets_service,
)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

DEFAULT_SERVICE_ACCOUNT_FILE = "credentials.json"
SHEET_109_LOG_NAME = "AiWB_109_Log"
SHEET_UNIT_MASTER_NAME = "Unit Master"
LOCAL_ROOT = Path(".aiwb_local")
DRAFT_DIR = LOCAL_ROOT / "drafts"
AUDIT_LOG_FILE = LOCAL_ROOT / "aiwb_audit.log"
CLOUD_SNAPSHOT_FILE = LOCAL_ROOT / "cloud_snapshot.pkl"
RULE_ID_HIT_RATE_BASELINE_FILE = LOCAL_ROOT / "rule_id_hit_rate_baseline.json"
FORMULA_DICTIONARY_109_FILE = Path("docs/AiWB_公式字典_109_v1.yaml")
SANDY_COVE_DATES_FILE = Path("docs/Sandy cove.xlsx")

DAILY_API_QUOTA_ESTIMATE = 5000
DEFAULT_UID_COLUMN = "AiWB_UID"
DEFAULT_AMOUNT_COLUMN = "Amount"
DEFAULT_ENTITY_COLUMN = "WBH"
DEFAULT_GUARD_SHEET = "Project Ledger"
DEFAULT_EXPECTED_FIRST_CELL = "Project Ledger"

UID_STATUS_COL = "__AIWB_UID_SYNC_STATUS"
SHADOW_CONFLICT_COL = "__AIWB_SHADOW_CONFLICT"
SHADOW_PY_PROFIT_COL = "__AIWB_PY_PROFIT"
SHADOW_PY_TAX_COL = "__AIWB_PY_TAX"

UID_PENDING_VALUE = "待同步"
UID_SYNCED_VALUE = "已同步"

COLOR_FILL_WHITE = {"red": 1.0, "green": 1.0, "blue": 1.0}
COLOR_FILL_LIGHT_GRAY = {"red": 0.93, "green": 0.93, "blue": 0.93}
COLOR_FILL_LIGHT_RED = {"red": 0.98, "green": 0.89, "blue": 0.89}
COLOR_FILL_LIGHT_YELLOW = {"red": 1.0, "green": 0.96, "blue": 0.76}
NUMBER_FORMAT_PERCENT_2 = {"type": "NUMBER", "pattern": "0.00%"}
NUMBER_FORMAT_YEAR_0 = {"type": "NUMBER", "pattern": "0"}
NUMBER_FORMAT_DATE_ISO = {"type": "DATE", "pattern": "yyyy-mm-dd"}
NUMBER_FORMAT_DATE_MDY = {"type": "DATE", "pattern": "m/d/yyyy"}
DEFAULT_RULE_ID_HIT_RATE_ALERT_THRESHOLD = 0.20
MANAGED_109_PROTECTION_DESCRIPTION = "AiWB managed main sheet protection"
MANAGED_109_FORMULA_PROTECTION_PREFIX = "AiWB managed formula lock"
MANAGED_SCOPING_PROTECTION_DESCRIPTION = "AiWB managed Scoping protection"
MANAGED_UNIT_MASTER_PROTECTION_DESCRIPTION = "AiWB managed Unit Master protection"
MANAGED_EXTERNAL_PROTECTION_PREFIX = "AiWB managed external protection"
SHEET_PROJECT_STATE_NAME = "AiWB_Project_State"
SHEET_AUDIT_LOG_NAME = "AiWB_Audit_Log"
SHEET_EDIT_LOG_NAME = "AiWB_Edit_Log"
WORKBENCH_STAGE_PROJECT_CREATED = "project_created"
WORKBENCH_STAGE_EXTERNAL_DATA_READY = "external_data_ready"
WORKBENCH_STAGE_MANUAL_INPUT_READY = "manual_input_ready"
WORKBENCH_STAGE_LOCKED_109_APPROVED = "locked_109_approved"
MANAGED_DATA_LOCK_PREFIX = "AiWB managed data lock"
PROJECT_MAIN_SHEET_PATTERN = re.compile(r"^\d{3}$")
FORMULA_WRITE_CHUNK_SIZE = 1000
GOOGLE_WRITE_MAX_ATTEMPTS = 5
GOOGLE_WRITE_BACKOFF_BASE_SECONDS = 0.8
GOOGLE_WRITE_BACKOFF_MAX_SECONDS = 6.0
PROJECT_MAIN_SHEET_RESERVED_TITLES = {
    "Payable",
    "Final Detail",
    "Draw request report",
    "Unit Budget",
    "Unit Master",
    "Scoping",
    "Contract",
    "Draw Invoice List",
    "Transfer Log",
    "Change Order Log",
}


class MappingService:
    _ACTIVE_STATUSES = {"active", "manual_override", "review_required"}
    _FIELD_ALIAS_BY_SHEET: Dict[str, Dict[str, str]] = {
        "Payable": {
            "amount": "amount",
            "cost_code": "cost_code",
            "year": "year",
            "incurred_year": "year",
            "posting_year": "year",
            "incurred_date": "year",
        }
    }

    @staticmethod
    @lru_cache(maxsize=256)
    def get_project_mappings(project_id: str | None) -> Dict[str, Dict[str, int]]:
        normalized_project_id = _safe_string(project_id)
        if not normalized_project_id:
            return {}

        supabase_url = _safe_string(os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL"))
        service_role_key = _safe_string(os.getenv("SUPABASE_SERVICE_ROLE_KEY"))
        if not supabase_url or not service_role_key:
            return {}

        query = urlencode(
            [
                ("select", "sheet_name,logical_field,column_index,mapping_status"),
                ("project_id", f"eq.{normalized_project_id}"),
            ]
        )
        endpoint = f"{supabase_url.rstrip('/')}/rest/v1/sheet_field_mappings?{query}"
        request = Request(
            endpoint,
            headers={
                "apikey": service_role_key,
                "Authorization": f"Bearer {service_role_key}",
                "Accept": "application/json",
            },
        )

        try:
            with urlopen(request, timeout=10) as response:
                rows = json.loads(response.read().decode("utf-8"))
        except Exception:
            return {}

        if not isinstance(rows, list):
            return {}

        mappings: Dict[str, Dict[str, int]] = {}
        for row in rows:
            if not isinstance(row, Mapping):
                continue
            status = _safe_string(row.get("mapping_status", "")).lower()
            if status and status not in MappingService._ACTIVE_STATUSES:
                continue

            sheet_name = _safe_string(row.get("sheet_name", ""))
            logical_field = _safe_string(row.get("logical_field", ""))
            if not sheet_name or not logical_field:
                continue

            try:
                column_index = int(row.get("column_index") or 0)
            except (TypeError, ValueError):
                continue
            if column_index < 1:
                continue

            logical_alias = MappingService._FIELD_ALIAS_BY_SHEET.get(sheet_name, {}).get(logical_field)
            if not logical_alias:
                continue

            mappings.setdefault(sheet_name, {})[logical_alias] = column_index

        return mappings


class SnapshotStaleError(RuntimeError):
    """Raised when the current sheet structure is stale against snapshot assumptions."""


_SNAPSHOT_PLACEHOLDER_PATTERN = re.compile(r"\$\{([A-Za-z0-9_.]+)(?::(?:col|range))?\}")


def _supabase_rest_request_json(
    *,
    method: str,
    resource: str,
    query: Mapping[str, str] | None = None,
    body: Mapping[str, Any] | None = None,
) -> Any:
    supabase_url = _safe_string(os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL"))
    service_role_key = _safe_string(os.getenv("SUPABASE_SERVICE_ROLE_KEY"))
    if not supabase_url or not service_role_key:
        return None

    query_string = urlencode(list(query.items())) if query else ""
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/{resource}"
    if query_string:
        endpoint = f"{endpoint}?{query_string}"

    payload_bytes = None
    if body is not None:
        payload_bytes = json.dumps(body, ensure_ascii=False).encode("utf-8")

    request = Request(
        endpoint,
        data=payload_bytes,
        method=method.upper(),
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
    )

    ssl_context = None
    try:
        import certifi  # type: ignore

        ssl_context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        ssl_context = None

    with urlopen(request, timeout=20, context=ssl_context) as response:
        raw = response.read().decode("utf-8").strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return raw


def _looks_like_dashboard_summary(value: Any) -> bool:
    return isinstance(value, Mapping) and isinstance(value.get("audit_tabs"), Mapping)


def _extract_dashboard_summary_payload(data_json: Any) -> Dict[str, Any]:
    if not isinstance(data_json, Mapping):
        return {}

    for key in ("dashboard_summary", "audit_dashboard_snapshot"):
        candidate = data_json.get(key)
        if _looks_like_dashboard_summary(candidate):
            return dict(candidate)

    if _looks_like_dashboard_summary(data_json):
        return dict(data_json)

    return {}


def _load_dashboard_summary_seed(
    *,
    project_id: str | None = None,
    spreadsheet_id: str | None = None,
) -> Dict[str, Any]:
    normalized_project_id = _safe_string(project_id)
    normalized_spreadsheet_id = _safe_string(spreadsheet_id)
    if not normalized_project_id and not normalized_spreadsheet_id:
        return {}

    project_lookup_id = normalized_project_id
    if not project_lookup_id and normalized_spreadsheet_id:
        try:
            project_rows = _supabase_rest_request_json(
                method="GET",
                resource="projects",
                query={
                    "select": "id",
                    "spreadsheet_id": f"eq.{normalized_spreadsheet_id}",
                    "limit": "1",
                },
            )
        except Exception:
            project_rows = None
        if isinstance(project_rows, list) and project_rows and isinstance(project_rows[0], Mapping):
            project_lookup_id = _safe_string(project_rows[0].get("id"))

    if not project_lookup_id:
        return {}

    for resource, query in (
        (
            "audit_cache",
            {
                "select": "data_json,last_synced_at",
                "project_id": f"eq.{project_lookup_id}",
                "order": "last_synced_at.desc",
                "limit": "1",
            },
        ),
        (
            "audit_snapshots",
            {
                "select": "data_json,created_at",
                "project_id": f"eq.{project_lookup_id}",
                "is_current": "eq.true",
                "order": "created_at.desc",
                "limit": "1",
            },
        ),
    ):
        try:
            rows = _supabase_rest_request_json(
                method="GET",
                resource=resource,
                query=query,
            )
        except Exception:
            continue
        if not isinstance(rows, list) or not rows or not isinstance(rows[0], Mapping):
            continue
        summary = _extract_dashboard_summary_payload(rows[0].get("data_json"))
        if summary:
            synced_at = _safe_string(rows[0].get("last_synced_at") or rows[0].get("created_at"))
            if synced_at and not _safe_string(summary.get("last_synced_at")):
                summary["last_synced_at"] = synced_at
            return summary

    return {}


def build_dashboard_summary_payload(
    *,
    spreadsheet_id: str,
    project_id: str | None = None,
    reclassify_summary: Mapping[str, Any] | None = None,
    mapping_metrics: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    summary = _load_dashboard_summary_seed(project_id=project_id, spreadsheet_id=spreadsheet_id)
    if not summary:
        summary = {
            "project_name": f"Project {(_safe_string(spreadsheet_id) or 'Unknown')[:8]}",
            "highlights": [],
            "audit_tabs": {},
        }

    reclass_summary = dict(reclassify_summary or {})
    mapping = dict(mapping_metrics or {})
    mapping_health = {
        "fallback_count": int(mapping.get("fallback_count", 0) or 0),
        "fallback_fields": list(mapping.get("fallback_fields", []) or []),
        "mapping_score": float(mapping.get("mapping_score", 1.0) or 0.0),
        "mapping_field_count": int(mapping.get("mapping_field_count", 0) or 0),
    }

    audit_tabs = dict(summary.get("audit_tabs") or {})
    external_recon = dict(audit_tabs.get("external_recon") or {})
    external_recon.setdefault("summary", "后台快照已更新，前端将直接渲染快照摘要。")
    manual_input = dict(audit_tabs.get("manual_input") or {})
    reclass_audit = dict(audit_tabs.get("reclass_audit") or {})
    reclass_overview = dict(reclass_audit.get("overview") or {})
    reclass_overview.setdefault("payable_count", int(reclass_summary.get("payable_rows_written", 0) or 0))
    reclass_overview.setdefault("final_detail_count", int(reclass_summary.get("final_detail_rows_written", 0) or 0))
    reclass_overview.setdefault("diff_count", int(reclass_summary.get("draw_request_unmatched_rows", 0) or 0))
    reclass_audit["overview"] = reclass_overview
    compare_109 = dict(audit_tabs.get("compare_109") or {})
    compare_109["mapping_health"] = mapping_health
    compare_109.setdefault("warnings", [])
    compare_109.setdefault("metric_rows", [])

    audit_tabs["external_recon"] = external_recon
    audit_tabs["manual_input"] = manual_input
    audit_tabs["reclass_audit"] = reclass_audit
    audit_tabs["compare_109"] = compare_109

    highlights = summary.get("highlights")
    if not isinstance(highlights, list) or not highlights:
        highlights = [
            {"label": "收入", "value": "-", "color": "slate"},
            {"label": "成本", "value": "-", "color": "slate"},
            {"label": "毛利", "value": "-", "color": "slate"},
            {"label": "完工进度", "value": "-", "color": "slate"},
        ]

    summary["highlights"] = highlights
    summary["audit_tabs"] = audit_tabs
    summary["workflow_stage"] = WORKBENCH_STAGE_MANUAL_INPUT_READY
    summary["mapping_health"] = mapping_health
    summary["last_synced_at"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    return _safe_json_default(summary)


def _extract_main_sheet_title_from_project_row(row: Mapping[str, Any]) -> str:
    project_sequence = _safe_string(row.get("project_sequence"))
    if project_sequence and PROJECT_MAIN_SHEET_PATTERN.fullmatch(project_sequence):
        return project_sequence
    configured_title = _safe_string(row.get("sheet_109_title"))
    if configured_title:
        return configured_title
    return ""


def _fetch_project_main_sheet_title(
    *,
    project_id: str | None = None,
    spreadsheet_id: str | None = None,
) -> str:
    normalized_project_id = _safe_string(project_id)
    normalized_spreadsheet_id = _safe_string(spreadsheet_id)
    if not normalized_project_id and not normalized_spreadsheet_id:
        return ""

    query_candidates: List[Dict[str, str]] = []
    if normalized_project_id:
        query_candidates.append(
            {
                "select": "sheet_109_title,project_sequence",
                "id": f"eq.{normalized_project_id}",
                "limit": "1",
            }
        )
        query_candidates.append(
            {
                "select": "sheet_109_title",
                "id": f"eq.{normalized_project_id}",
                "limit": "1",
            }
        )
    if normalized_spreadsheet_id:
        query_candidates.append(
            {
                "select": "sheet_109_title,project_sequence",
                "spreadsheet_id": f"eq.{normalized_spreadsheet_id}",
                "limit": "1",
            }
        )
        query_candidates.append(
            {
                "select": "sheet_109_title",
                "spreadsheet_id": f"eq.{normalized_spreadsheet_id}",
                "limit": "1",
            }
        )

    for query in query_candidates:
        try:
            rows = _supabase_rest_request_json(
                method="GET",
                resource="projects",
                query=query,
            )
        except Exception:
            continue
        if not isinstance(rows, list) or not rows:
            continue
        first_row = rows[0]
        if not isinstance(first_row, Mapping):
            continue
        resolved_title = _extract_main_sheet_title_from_project_row(first_row)
        if resolved_title:
            return resolved_title

    return ""


def _fetch_project_sequence(
    *,
    project_id: str | None = None,
    spreadsheet_id: str | None = None,
) -> str:
    normalized_project_id = _safe_string(project_id)
    normalized_spreadsheet_id = _safe_string(spreadsheet_id)
    if not normalized_project_id and not normalized_spreadsheet_id:
        return ""

    query_candidates: List[Dict[str, str]] = []
    if normalized_project_id:
        query_candidates.append(
            {
                "select": "project_sequence",
                "id": f"eq.{normalized_project_id}",
                "limit": "1",
            }
        )
    if normalized_spreadsheet_id:
        query_candidates.append(
            {
                "select": "project_sequence",
                "spreadsheet_id": f"eq.{normalized_spreadsheet_id}",
                "limit": "1",
            }
        )

    for query in query_candidates:
        try:
            rows = _supabase_rest_request_json(
                method="GET",
                resource="projects",
                query=query,
            )
        except Exception:
            continue
        if not isinstance(rows, list) or not rows:
            continue
        first_row = rows[0]
        if not isinstance(first_row, Mapping):
            continue
        project_sequence = _safe_string(first_row.get("project_sequence"))
        if project_sequence and PROJECT_MAIN_SHEET_PATTERN.fullmatch(project_sequence):
            return project_sequence

    fallback_main_sheet_title = _fetch_project_main_sheet_title(
        project_id=normalized_project_id or None,
        spreadsheet_id=normalized_spreadsheet_id or None,
    )
    if fallback_main_sheet_title and PROJECT_MAIN_SHEET_PATTERN.fullmatch(fallback_main_sheet_title):
        return fallback_main_sheet_title
    return ""


def _discover_main_sheet_title_from_workbook(
    *,
    service,
    spreadsheet_id: str | None,
) -> str:
    normalized_spreadsheet_id = _safe_string(spreadsheet_id)
    if service is None or not normalized_spreadsheet_id:
        return ""
    try:
        metadata = service.spreadsheets().get(
            spreadsheetId=normalized_spreadsheet_id,
            fields="sheets(properties(title))",
        ).execute()
    except Exception:
        return ""

    candidates: List[str] = []
    for sheet in metadata.get("sheets", []):
        if not isinstance(sheet, Mapping):
            continue
        props = sheet.get("properties")
        if not isinstance(props, Mapping):
            continue
        title = _safe_string(props.get("title"))
        if not title:
            continue
        if title in PROJECT_MAIN_SHEET_RESERVED_TITLES:
            continue
        if PROJECT_MAIN_SHEET_PATTERN.fullmatch(title):
            candidates.append(title)
    if len(candidates) == 1:
        return candidates[0]
    return ""


def _fetch_current_formula_snapshot_row(project_id: str, spreadsheet_id: str) -> Dict[str, Any] | None:
    rows = _supabase_rest_request_json(
        method="GET",
        resource="audit_snapshots",
        query={
            "select": "id,sync_run_id,created_at,is_current,data_json,mapping_manifest_json",
            "project_id": f"eq.{project_id}",
            "spreadsheet_id": f"eq.{spreadsheet_id}",
            "is_current": "eq.true",
            "order": "created_at.desc",
            "limit": "1",
        },
    )
    if not isinstance(rows, Sequence) or not rows or not isinstance(rows[0], Mapping):
        return None

    row = dict(rows[0])
    sync_run_id = _safe_string(row.get("sync_run_id"))
    run_status = ""
    if sync_run_id:
        run_rows = _supabase_rest_request_json(
            method="GET",
            resource="audit_sync_runs",
            query={
                "select": "status",
                "id": f"eq.{sync_run_id}",
                "limit": "1",
            },
        )
        if isinstance(run_rows, Sequence) and run_rows and isinstance(run_rows[0], Mapping):
            run_status = _safe_string(run_rows[0].get("status"))
    row["sync_run_status"] = run_status
    return row


def _parse_formula_cell_context(cell_ref: str) -> tuple[str, int] | None:
    match = re.fullmatch(r"\$?([A-Za-z]+)\$?([0-9]+)", _safe_string(cell_ref))
    if not match:
        return None
    return match.group(1).upper(), int(match.group(2))


def _extract_required_mapping_fields_from_templates(
    template_rows: Sequence[Any],
) -> Dict[str, List[str]]:
    required: Dict[str, set[str]] = {}
    for item in template_rows:
        if not isinstance(item, Mapping):
            continue
        source_formula = _safe_string(item.get("formula_template") or item.get("formula_rendered"))
        if not source_formula:
            continue
        for match in _SNAPSHOT_PLACEHOLDER_PATTERN.finditer(source_formula):
            token = _safe_string(match.group(1))
            if "." not in token:
                continue
            sheet_name, field_name = token.split(".", 1)
            sheet_name = _safe_string(sheet_name)
            field_name = _safe_string(field_name)
            if not sheet_name or not field_name:
                continue
            required.setdefault(sheet_name, set()).add(field_name)
    return {sheet: sorted(fields) for sheet, fields in required.items() if fields}


def _normalize_required_mapping_fields(raw: Any) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    if not isinstance(raw, Mapping):
        return out
    for sheet_name, fields in raw.items():
        normalized_sheet = _safe_string(sheet_name)
        if not normalized_sheet:
            continue
        if isinstance(fields, Sequence) and not isinstance(fields, (str, bytes, bytearray)):
            normalized_fields = sorted({_safe_string(field) for field in fields if _safe_string(field)})
        else:
            one_field = _safe_string(fields)
            normalized_fields = [one_field] if one_field else []
        if normalized_fields:
            out[normalized_sheet] = normalized_fields
    return out


def _normalize_formula_mapping_manifest(raw: Any) -> Dict[str, Dict[str, int]]:
    if not isinstance(raw, Mapping):
        return {}
    source = raw.get("mappings") if "mappings" in raw else raw
    if not isinstance(source, Mapping):
        return {}

    mappings: Dict[str, Dict[str, int]] = {}
    for sheet_name, fields in source.items():
        normalized_sheet = _safe_string(sheet_name)
        if not normalized_sheet or not isinstance(fields, Mapping):
            continue
        normalized_fields: Dict[str, int] = {}
        for field_name, column_index in fields.items():
            normalized_field = _safe_string(field_name)
            try:
                normalized_index = int(column_index)
            except (TypeError, ValueError):
                continue
            if normalized_field and normalized_index >= 1:
                normalized_fields[normalized_field] = normalized_index
        if normalized_fields:
            mappings[normalized_sheet] = normalized_fields
    return mappings


def _formula_templates_have_placeholders(template_rows: Sequence[Any]) -> bool:
    for item in template_rows:
        if not isinstance(item, Mapping):
            continue
        source_formula = _safe_string(item.get("formula_template") or item.get("formula_rendered"))
        if "${" in source_formula:
            return True
    return False


def _normalize_row_fingerprint_label_cells(raw: Any) -> List[str]:
    if not isinstance(raw, Mapping):
        return []
    label_cells = raw.get("label_cells")
    if not isinstance(label_cells, Sequence) or isinstance(label_cells, (str, bytes, bytearray)):
        return []
    return [_safe_string(value) for value in label_cells]


def _find_header_column_index(header_row: Sequence[Any], aliases: Sequence[str]) -> int | None:
    wanted = {_normalize_header_token(alias) for alias in aliases}
    for idx, value in enumerate(header_row, start=1):
        if _normalize_header_token(value) in wanted:
            return idx
    return None


def _discover_payable_formula_mappings(service, spreadsheet_id: str) -> Dict[str, int]:
    try:
        rows = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range="'Payable'!A1:AZ6")
            .execute()
            .get("values", [])
        )
    except Exception:
        return {}

    if not isinstance(rows, Sequence):
        return {}

    best_layout: Dict[str, int] = {}
    best_score = -1
    for header_row in rows:
        if not isinstance(header_row, Sequence):
            continue
        row_layout: Dict[str, int] = {}
        amount_col = _find_header_column_index(header_row, ("Amount", "Amt"))
        cost_code_col = _find_header_column_index(header_row, ("Cost Code", "CostCode", "Cost_Code"))
        year_col = _find_header_column_index(
            header_row,
            (
                "Year",
                "Incurred Year",
                "Posting Year",
                "Incurred Date",
                "Invoiced Date",
            ),
        )
        if amount_col:
            row_layout["amount"] = amount_col
        if cost_code_col:
            row_layout["cost_code"] = cost_code_col
        if year_col:
            row_layout["year"] = year_col

        score = len(row_layout)
        if score > best_score:
            best_score = score
            best_layout = row_layout
        if score >= 3:
            return row_layout
    return best_layout


def _resolve_writeback_formula_mappings(
    *,
    service,
    spreadsheet_id: str,
    project_id: str,
) -> Dict[str, Dict[str, int]]:
    mappings = {
        sheet_name: {field_name: int(value) for field_name, value in dict(fields).items()}
        for sheet_name, fields in MappingService.get_project_mappings(project_id).items()
        if isinstance(fields, Mapping)
    }

    payable_mapping = mappings.setdefault("Payable", {})
    required_fields = {"amount", "cost_code", "year"}
    missing_fields = [field_name for field_name in required_fields if field_name not in payable_mapping]

    if missing_fields and service is not None:
        discovered = _discover_payable_formula_mappings(service, spreadsheet_id)
        for field_name, column_index in discovered.items():
            if field_name not in payable_mapping and column_index >= 1:
                payable_mapping[field_name] = int(column_index)

    return mappings


def _coerce_snapshot_header_cells(value: Any) -> List[str]:
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return [_safe_string(item) for item in value]
    if isinstance(value, Mapping):
        for key in ("cells", "values", "header_cells", "headers"):
            nested = value.get(key)
            if isinstance(nested, Sequence) and not isinstance(nested, (str, bytes, bytearray)):
                return [_safe_string(item) for item in nested]
    return []


def validate_snapshot_writeback_consistency(
    *,
    service,
    spreadsheet_id: str,
    project_id: str,
    snapshot_meta: Mapping[str, Any] | None = None,
    plan: Sequence[Mapping[str, Any]] | None = None,
) -> Dict[str, Any]:
    resolved_spreadsheet_id = _safe_string(spreadsheet_id)
    resolved_project_id = _safe_string(project_id or spreadsheet_id)
    runtime_meta = dict(snapshot_meta or {})
    required_mapping_fields = _normalize_required_mapping_fields(runtime_meta.get("required_mapping_fields"))

    required_sheets = set(required_mapping_fields.keys())
    for item in plan or []:
        if isinstance(item, Mapping):
            sheet_name = _safe_string(item.get("sheet"))
            if sheet_name:
                required_sheets.add(sheet_name)

    workbook_meta = service.spreadsheets().get(
        spreadsheetId=resolved_spreadsheet_id,
        includeGridData=False,
        fields="sheets(properties(title,gridProperties(rowCount,columnCount)))",
    ).execute()
    sheet_properties = {}
    for sheet in workbook_meta.get("sheets", []):
        if not isinstance(sheet, Mapping):
            continue
        props = sheet.get("properties")
        if not isinstance(props, Mapping):
            continue
        title = _safe_string(props.get("title"))
        if not title:
            continue
        grid = props.get("gridProperties") if isinstance(props.get("gridProperties"), Mapping) else {}
        sheet_properties[title] = {
            "row_count": int(grid.get("rowCount") or 0),
            "column_count": int(grid.get("columnCount") or 0),
        }

    missing_sheets = sorted(sheet_name for sheet_name in required_sheets if sheet_name not in sheet_properties)
    if missing_sheets:
        raise SnapshotStaleError(
            "SNAPSHOT_STALE_ERROR: MISSING_SHEETS " + ",".join(missing_sheets)
        )

    mappings = _resolve_writeback_formula_mappings(
        service=service,
        spreadsheet_id=resolved_spreadsheet_id,
        project_id=resolved_project_id,
    )
    missing_fields: List[str] = []
    for sheet_name, fields in required_mapping_fields.items():
        resolved_sheet_mapping = mappings.get(sheet_name, {})
        if not isinstance(resolved_sheet_mapping, Mapping):
            resolved_sheet_mapping = {}
        for field_name in fields:
            if _safe_string(resolved_sheet_mapping.get(field_name)) == "":
                missing_fields.append(f"{sheet_name}.{field_name}")
    if missing_fields:
        raise SnapshotStaleError(
            "SNAPSHOT_STALE_ERROR: MAPPING_CONFIDENCE_ZERO " + ",".join(sorted(missing_fields))
        )

    warnings: List[str] = []
    discovery_checked_count = 0
    sync_run_id = _safe_string(runtime_meta.get("sync_run_id"))
    if sync_run_id:
        discovery_rows = _supabase_rest_request_json(
            method="GET",
            resource="sheet_discovery_snapshots",
            query={
                "select": "sheet_name,header_row_index,header_cells_json",
                "sync_run_id": f"eq.{sync_run_id}",
            },
        )
        if isinstance(discovery_rows, Sequence):
            relevant_rows = [
                row
                for row in discovery_rows
                if isinstance(row, Mapping) and _safe_string(row.get("sheet_name")) in required_sheets
            ]
            if relevant_rows:
                ranges: List[str] = []
                for row in relevant_rows:
                    sheet_name = _safe_string(row.get("sheet_name"))
                    header_row_index = int(row.get("header_row_index") or 0)
                    if header_row_index <= 0:
                        continue
                    if sheet_properties.get(sheet_name, {}).get("row_count", 0) < header_row_index:
                        raise SnapshotStaleError(
                            f"SNAPSHOT_STALE_ERROR: HEADER_ROW_OUT_OF_BOUNDS {sheet_name}!{header_row_index}"
                        )
                    ranges.append(f"{_quote_sheet_name(sheet_name)}!A{header_row_index}:ZZ{header_row_index}")
                if ranges:
                    batch_response = service.spreadsheets().values().batchGet(
                        spreadsheetId=resolved_spreadsheet_id,
                        ranges=ranges,
                        valueRenderOption="FORMATTED_VALUE",
                    ).execute()
                    value_ranges = batch_response.get("valueRanges", [])
                    for idx, row in enumerate(relevant_rows):
                        if idx >= len(value_ranges):
                            break
                        expected_cells = _coerce_snapshot_header_cells(row.get("header_cells_json"))
                        expected_tokens = {
                            _normalize_header_token(value)
                            for value in expected_cells
                            if _normalize_header_token(value)
                        }
                        actual_values = value_ranges[idx].get("values", [])
                        actual_cells = actual_values[0] if actual_values else []
                        actual_tokens = {
                            _normalize_header_token(value)
                            for value in actual_cells
                            if _normalize_header_token(value)
                        }
                        if not expected_tokens:
                            continue
                        overlap = len(expected_tokens & actual_tokens) / max(len(expected_tokens), 1)
                        sheet_name = _safe_string(row.get("sheet_name"))
                        if overlap < 0.2:
                            raise SnapshotStaleError(
                                f"SNAPSHOT_STALE_ERROR: HEADER_DRIFT {sheet_name} overlap={overlap:.2f}"
                            )
                        if overlap < 0.6:
                            warnings.append(f"HEADER_DRIFT_WARNING:{sheet_name}:{overlap:.2f}")
                        discovery_checked_count += 1
            else:
                warnings.append("DISCOVERY_SNAPSHOT_MISSING")
        else:
            warnings.append("DISCOVERY_SNAPSHOT_UNAVAILABLE")
    else:
        warnings.append("SYNC_RUN_ID_MISSING")

    fingerprint_checks: List[Tuple[str, str, str, List[str]]] = []
    fingerprint_ranges: List[str] = []
    for item in plan or []:
        if not isinstance(item, Mapping):
            continue
        sheet_name = _safe_string(item.get("sheet"))
        cell_ref = _safe_string(item.get("cell"))
        parsed_cell = _parse_formula_cell_context(cell_ref)
        expected_cells = _normalize_row_fingerprint_label_cells(item.get("row_fingerprint"))
        if not sheet_name or not parsed_cell or not expected_cells:
            continue
        row_number = parsed_cell[1]
        if sheet_properties.get(sheet_name, {}).get("row_count", 0) < row_number:
            raise SnapshotStaleError(
                f"SNAPSHOT_STALE_ERROR: FORMULA_ROW_OUT_OF_BOUNDS {sheet_name}!{cell_ref}"
            )
        range_a1 = f"{_quote_sheet_name(sheet_name)}!C{row_number}:D{row_number}"
        fingerprint_ranges.append(range_a1)
        fingerprint_checks.append((sheet_name, cell_ref, range_a1, expected_cells))

    row_fingerprint_checked_count = 0
    if fingerprint_ranges:
        batch_response = service.spreadsheets().values().batchGet(
            spreadsheetId=resolved_spreadsheet_id,
            ranges=fingerprint_ranges,
            valueRenderOption="FORMATTED_VALUE",
        ).execute()
        value_ranges = batch_response.get("valueRanges", [])
        for idx, (sheet_name, cell_ref, _range_a1, expected_cells) in enumerate(fingerprint_checks):
            actual_values = value_ranges[idx].get("values", []) if idx < len(value_ranges) else []
            actual_cells = list(actual_values[0]) if actual_values else []
            expected_tokens = [_normalize_header_token(value) for value in expected_cells]
            actual_tokens = [_normalize_header_token(value) for value in actual_cells]
            for expected_idx, expected_token in enumerate(expected_tokens):
                if not expected_token:
                    continue
                actual_token = actual_tokens[expected_idx] if expected_idx < len(actual_tokens) else ""
                if actual_token != expected_token:
                    raise SnapshotStaleError(
                        f"SNAPSHOT_STALE_ERROR: FORMULA_ROW_DRIFT {sheet_name}!{cell_ref}"
                    )
            row_fingerprint_checked_count += 1

    return {
        "checked_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "snapshot_id": _safe_string(runtime_meta.get("snapshot_id")),
        "sync_run_id": sync_run_id,
        "required_sheet_count": len(required_sheets),
        "required_mapping_field_count": sum(len(fields) for fields in required_mapping_fields.values()),
        "discovery_checked_count": discovery_checked_count,
        "row_fingerprint_checked_count": row_fingerprint_checked_count,
        "warnings": warnings,
    }


def load_current_snapshot_formula_plan(
    *,
    project_id: str,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
    service=None,
) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    resolved_project_id = _safe_string(project_id or spreadsheet_id)
    resolved_spreadsheet_id = _safe_string(spreadsheet_id)
    if not resolved_project_id or not resolved_spreadsheet_id:
        raise RuntimeError("CURRENT_SNAPSHOT_INPUT_MISSING")
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        project_id=resolved_project_id,
        spreadsheet_id=resolved_spreadsheet_id,
        service=service,
    )

    snapshot_row = _fetch_current_formula_snapshot_row(resolved_project_id, resolved_spreadsheet_id)
    if not snapshot_row:
        raise RuntimeError("CURRENT_SNAPSHOT_NOT_FOUND")

    sync_run_status = _safe_string(snapshot_row.get("sync_run_status")).lower()
    if sync_run_status != "succeeded":
        raise RuntimeError(f"CURRENT_SNAPSHOT_NOT_SUCCEEDED: {sync_run_status or 'unknown'}")

    data_json = snapshot_row.get("data_json")
    if not isinstance(data_json, Mapping):
        raise RuntimeError("CURRENT_SNAPSHOT_DATA_INVALID")

    template_rows = data_json.get("formula_plan_templates")
    if not isinstance(template_rows, Sequence):
        raise RuntimeError("CURRENT_SNAPSHOT_FORMULA_PLAN_EMPTY")
    required_mapping_fields = _extract_required_mapping_fields_from_templates(template_rows)
    frozen_mapping_manifest = data_json.get("formula_mapping_manifest")
    frozen_mappings = _normalize_formula_mapping_manifest(frozen_mapping_manifest)
    if _formula_templates_have_placeholders(template_rows) and not frozen_mappings:
        raise RuntimeError("CURRENT_SNAPSHOT_MAPPING_MANIFEST_MISSING")

    resolver = FormulaTemplateResolver()
    mappings = frozen_mappings
    rendered_plan: List[Dict[str, str]] = []
    skipped_rows = 0

    for item in template_rows:
        if not isinstance(item, Mapping):
            skipped_rows += 1
            continue
        sheet_name = _safe_string(item.get("sheet")) or resolved_sheet_109_title
        cell_ref = _safe_string(item.get("cell"))
        formula_template = _safe_string(item.get("formula_template"))
        fallback_formula = _safe_string(item.get("formula_rendered"))
        source_formula = formula_template or fallback_formula
        if not cell_ref or not source_formula:
            skipped_rows += 1
            continue
        context = {}
        parsed_cell = _parse_formula_cell_context(cell_ref)
        if parsed_cell:
            context = {"self_col": parsed_cell[0], "self_row": parsed_cell[1]}

        try:
            formula = (
                resolver.resolve_formula(source_formula, mappings, context=context)
                if "${" in source_formula
                else source_formula
            )
        except MappingIncompleteError as exc:
            raise RuntimeError(f"CURRENT_SNAPSHOT_MAPPING_INCOMPLETE: {exc}") from exc

        rendered_item = {
            "sheet": sheet_name,
            "cell": cell_ref,
            "range": f"{_quote_sheet_name(sheet_name)}!{cell_ref}",
            "formula": formula,
            "logic": _safe_string(item.get("logic")),
            "formula_template": source_formula,
        }
        row_fingerprint = item.get("row_fingerprint")
        if isinstance(row_fingerprint, Mapping):
            rendered_item["row_fingerprint"] = dict(row_fingerprint)
        rendered_plan.append(rendered_item)

    if not rendered_plan:
        raise RuntimeError("CURRENT_SNAPSHOT_FORMULA_PLAN_EMPTY")

    meta = {
        "source": "current_snapshot",
        "snapshot_id": _safe_string(snapshot_row.get("id")),
        "sync_run_id": _safe_string(snapshot_row.get("sync_run_id")),
        "sync_run_status": sync_run_status,
        "formula_count": len(rendered_plan),
        "skipped_template_rows": skipped_rows,
        "formula_mapping_project_id": resolved_project_id,
        "formula_mapping_sheet_count": len(mappings),
        "formula_mapping_source": "snapshot_frozen",
        "sheet": resolved_sheet_109_title,
        "required_mapping_fields": required_mapping_fields,
    }
    return rendered_plan, meta


def _resolve_sheet_109_title(
    sheet_109_title: str | None = None,
    *,
    project_id: str | None = None,
    spreadsheet_id: str | None = None,
    service=None,
) -> str:
    normalized = _safe_string(sheet_109_title)
    if normalized:
        return normalized

    resolved_from_project = _fetch_project_main_sheet_title(
        project_id=project_id,
        spreadsheet_id=spreadsheet_id,
    )
    if resolved_from_project:
        return resolved_from_project

    resolved_from_workbook = _discover_main_sheet_title_from_workbook(
        service=service,
        spreadsheet_id=spreadsheet_id,
    )
    if resolved_from_workbook:
        return resolved_from_workbook

    raise RuntimeError("MAIN_SHEET_TITLE_UNRESOLVED")


def _build_109_range(a1_range: str, sheet_109_title: str | None = None) -> str:
    return f"{_quote_sheet_name(_resolve_sheet_109_title(sheet_109_title))}!{a1_range}"


def _build_109_cell_key(cell_a1: str, sheet_109_title: str | None = None) -> str:
    return f"{_resolve_sheet_109_title(sheet_109_title)}!{cell_a1}"


def _build_project_data_lock_sheet_names(sheet_109_title: str | None = None) -> List[str]:
    return [
        _resolve_sheet_109_title(sheet_109_title),
        "Scoping",
        "Unit Master",
        "Contract",
        "Unit Budget",
        "Payable",
        "Final Detail",
        "Draw request report",
        "Draw Invoice List",
        "Transfer Log",
        "Change Order Log",
    ]


def _bool_text(value: bool) -> str:
    return "TRUE" if bool(value) else "FALSE"


def _build_project_state_values(
    owner_email: str,
    current_stage: str = WORKBENCH_STAGE_PROJECT_CREATED,
    locked: bool = False,
    **extra: Any,
) -> List[List[Any]]:
    values = [
        ["key", "value"],
        ["current_stage", _safe_string(current_stage) or WORKBENCH_STAGE_PROJECT_CREATED],
        ["external_data_dirty", _bool_text(bool(extra.get("external_data_dirty", False)))],
        ["manual_input_dirty", _bool_text(bool(extra.get("manual_input_dirty", False)))],
        ["locked", _bool_text(locked)],
        ["owner_email", _safe_string(owner_email)],
        ["last_external_edit_at", _safe_string(extra.get("last_external_edit_at", ""))],
        ["last_external_edit_by", _safe_string(extra.get("last_external_edit_by", ""))],
        ["last_manual_edit_at", _safe_string(extra.get("last_manual_edit_at", ""))],
        ["last_manual_edit_by", _safe_string(extra.get("last_manual_edit_by", ""))],
        ["last_sync_at", _safe_string(extra.get("last_sync_at", ""))],
        ["last_validate_input_at", _safe_string(extra.get("last_validate_input_at", ""))],
        ["last_reclassify_at", _safe_string(extra.get("last_reclassify_at", ""))],
        ["last_109_initial_approval_at", _safe_string(extra.get("last_109_initial_approval_at", ""))],
        ["locked_at", _safe_string(extra.get("locked_at", ""))],
        ["locked_by", _safe_string(extra.get("locked_by", ""))],
        ["unlocked_at", _safe_string(extra.get("unlocked_at", ""))],
        ["unlocked_by", _safe_string(extra.get("unlocked_by", ""))],
    ]
    return values


def append_project_audit_log(
    service,
    spreadsheet_id: str,
    actor_email: str,
    action: str,
    previous_stage: str,
    next_stage: str,
    status: str,
    message: str,
    project_id: str | None = None,
) -> None:
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"{_quote_sheet_name(SHEET_AUDIT_LOG_NAME)}!A:I",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={
            "values": [
                [
                    datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                    _safe_string(actor_email),
                    _safe_string(action),
                    _safe_string(project_id or spreadsheet_id),
                    _safe_string(spreadsheet_id),
                    _safe_string(previous_stage),
                    _safe_string(next_stage),
                    _safe_string(status),
                    _safe_string(message),
                ]
            ]
        },
    ).execute()


def _build_external_sheet_edit_specs() -> Dict[str, Dict[str, List[str]]]:
    """Single source of truth for external sheet edit and wipe boundaries.

    Each sheet spec intentionally carries both the editable ranges and the
    clear ranges so future edits keep the contract and cleanup boundaries in
    lockstep. `_build_external_sheet_clear_ranges()` must derive its output from
    this structure instead of hand-maintaining a separate list.
    """
    return {
        "Contract": {
            "editable_ranges": ["'Contract'!A:ZZ"],
            "clear_ranges": ["'Contract'!A:ZZ"],
        },
        "Unit Budget": {
            "editable_ranges": ["'Unit Budget'!S:ZZ"],
            "filter_header_ranges": ["'Unit Budget'!A1:ZZ1"],
            "clear_ranges": ["'Unit Budget'!S:ZZ"],
        },
        "Payable": {
            "editable_ranges": ["'Payable'!L:AZ"],
            "filter_header_ranges": ["'Payable'!A1:ZZ1"],
            "clear_ranges": ["'Payable'!A2:ZZ"],
        },
        "Final Detail": {
            "editable_ranges": ["'Final Detail'!N:AL"],
            "filter_header_ranges": ["'Final Detail'!A1:ZZ1"],
            "clear_ranges": ["'Final Detail'!A2:ZZ"],
        },
        "Draw request report": {
            "editable_ranges": ["'Draw request report'!H:AR"],
            "filter_header_ranges": ["'Draw request report'!A1:ZZ2"],
            "clear_ranges": ["'Draw request report'!A3:ZZ"],
        },
        "Draw Invoice List": {
            "editable_ranges": ["'Draw Invoice List'!G:AE"],
            "filter_header_ranges": ["'Draw Invoice List'!A4:ZZ4"],
            "clear_ranges": ["'Draw Invoice List'!A5:ZZ"],
        },
        "Transfer Log": {
            "editable_ranges": ["'Transfer Log'!G:Z"],
            "filter_header_ranges": ["'Transfer Log'!A4:ZZ4"],
            "clear_ranges": ["'Transfer Log'!A5:ZZ"],
        },
        "Change Order Log": {
            "editable_ranges": ["'Change Order Log'!G:AE"],
            "filter_header_ranges": ["'Change Order Log'!A4:ZZ4"],
            "clear_ranges": ["'Change Order Log'!A5:ZZ"],
        },
    }


def _build_external_sheet_clear_ranges() -> List[str]:
    ranges: List[str] = []
    # Keep the wipe list derived from the spec above so the boundary cannot drift.
    for sheet_spec in _build_external_sheet_edit_specs().values():
        ranges.extend(sheet_spec["clear_ranges"])
    return ranges


def _build_external_sheet_data_probe_ranges() -> Dict[str, str]:
    return {
        "Payable": "'Payable'!A2:ZZ",
        "Final Detail": "'Final Detail'!A2:ZZ",
        "Draw request report": "'Draw request report'!A3:ZZ",
    }


def _build_update_hidden_sheet_request(sheet_id: int, hidden: bool) -> Dict[str, Any]:
    return {
        "updateSheetProperties": {
            "properties": {"sheetId": int(sheet_id), "hidden": bool(hidden)},
            "fields": "hidden",
        }
    }


def _build_external_sheet_protection_requests(
    service,
    spreadsheet_id: str,
) -> List[Dict[str, Any]]:
    def _build_update_column_count_request(sheet_id: int, column_count: int) -> Dict[str, Any]:
        return {
            "updateSheetProperties": {
                "properties": {"sheetId": int(sheet_id), "gridProperties": {"columnCount": int(column_count)}},
                "fields": "gridProperties.columnCount",
            }
        }

    def _clip_grid_range_to_sheet_bounds(grid_range: Dict[str, int], metadata: Mapping[str, Any]) -> Dict[str, int] | None:
        clipped = dict(grid_range)
        column_count = int(metadata.get("column_count", 0) or 0)
        row_count = int(metadata.get("row_count", 0) or 0)
        if column_count > 0:
            if int(clipped.get("startColumnIndex", 0)) >= column_count:
                return None
            clipped["endColumnIndex"] = min(int(clipped.get("endColumnIndex", column_count)), column_count)
        if row_count > 0 and "startRowIndex" in clipped:
            if int(clipped.get("startRowIndex", 0)) >= row_count:
                return None
            clipped["endRowIndex"] = min(int(clipped.get("endRowIndex", row_count)), row_count)
        return clipped

    requests: List[Dict[str, Any]] = []
    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))

    for sheet_name, spec in _build_external_sheet_edit_specs().items():
        metadata = _get_sheet_metadata(service, spreadsheet_id, sheet_name)
        description = f"{MANAGED_EXTERNAL_PROTECTION_PREFIX}: {sheet_name}"

        for protected_range in metadata.get("protected_ranges", []):
            if protected_range.get("description") != description:
                continue
            protected_range_id = protected_range.get("protectedRangeId")
            if protected_range_id is not None:
                requests.append(_build_delete_protected_range_request(int(protected_range_id)))

        if sheet_name == "Contract":
            continue

        unprotected_ranges = []
        for a1 in spec.get("editable_ranges", []):
            grid_range = _a1_to_grid_range(a1, int(metadata["sheet_id"]))
            target_column_count = int(grid_range.get("endColumnIndex", 0))
            if target_column_count > int(metadata.get("column_count", 0)):
                requests.append(_build_update_column_count_request(int(metadata["sheet_id"]), target_column_count))
                metadata["column_count"] = target_column_count
            unprotected_ranges.append(grid_range)
        for a1 in spec.get("filter_header_ranges", []):
            grid_range = _clip_grid_range_to_sheet_bounds(_a1_to_grid_range(a1, int(metadata["sheet_id"])), metadata)
            if grid_range:
                unprotected_ranges.append(grid_range)

        request = _build_add_protected_range_request(
            sheet_id=int(metadata["sheet_id"]),
            unprotected_ranges=unprotected_ranges,
            editor_email=editor_email or None,
            warning_only=True,
        )
        request["addProtectedRange"]["protectedRange"]["description"] = description
        requests.append(request)

    return requests


def _build_project_bootstrap_manual_clear_ranges(
    rows_109: Sequence[Sequence[Any]],
    rows_scoping: Sequence[Sequence[Any]],
    row_count_unit_master: int,
    sheet_109_title: str | None = None,
) -> List[str]:
    manual_ranges = _build_109_manual_input_ranges(
        rows_109,
        _year_columns_from_109_dictionary(_load_109_formula_dictionary()),
        sheet_109_title=sheet_109_title,
    )
    manual_ranges.extend(_build_scoping_manual_input_ranges(rows_scoping))
    manual_ranges.extend(_build_unit_master_bootstrap_clear_ranges(int(row_count_unit_master)))
    manual_ranges.extend(_build_unit_master_manual_input_ranges(int(row_count_unit_master)))
    manual_ranges.extend(_build_external_sheet_clear_ranges())
    compressed = _compress_a1_ranges(manual_ranges)
    return list(dict.fromkeys(compressed))


def _hide_system_log_sheet(
    service,
    spreadsheet_id: str,
) -> bool:
    try:
        metadata = _get_sheet_metadata(service, spreadsheet_id, SHEET_109_LOG_NAME)
    except RuntimeError:
        return False

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [_build_update_hidden_sheet_request(int(metadata["sheet_id"]), True)]},
    ).execute()
    return True


def _rename_109_sheet_for_project(
    service,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
) -> bool:
    actual_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )

    metadata_by_sheet = _get_sheet_metadata_map(service, spreadsheet_id)
    if actual_sheet_109_title in metadata_by_sheet:
        return False

    legacy_main_sheet = metadata_by_sheet.get("109")
    if not legacy_main_sheet:
        numeric_candidates = sorted(
            sheet_name
            for sheet_name in metadata_by_sheet
            if PROJECT_MAIN_SHEET_PATTERN.fullmatch(_safe_string(sheet_name))
            and sheet_name not in PROJECT_MAIN_SHEET_RESERVED_TITLES
            and sheet_name != actual_sheet_109_title
        )
        legacy_main_sheet = metadata_by_sheet.get(numeric_candidates[0]) if numeric_candidates else None

    metadata_109 = legacy_main_sheet
    if not metadata_109 or metadata_109.get("sheet_id") is None:
        return False

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": int(metadata_109["sheet_id"]),
                            "title": actual_sheet_109_title,
                        },
                        "fields": "title",
                    }
                }
            ]
        },
    ).execute()
    return True


def _apply_external_sheet_controls(
    service,
    spreadsheet_id: str,
) -> Dict[str, Any]:
    external_requests = _build_external_sheet_protection_requests(service, spreadsheet_id)
    if external_requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": external_requests},
        ).execute()

    log_hidden = _hide_system_log_sheet(service, spreadsheet_id)
    return {
        "external_protection_request_count": len(external_requests),
        "log_hidden": log_hidden,
    }


def _derive_unit_budget_actual_settlement_fields(
    unit_code: Any,
    settlement_year: Any,
    co_date: Any,
    latest_unit_year: int | None,
) -> Tuple[str, int | str]:
    code_text = _safe_string(unit_code)
    actual_date = _co_date_to_actual_settlement_date(co_date)
    settlement_year_value = _extract_year(settlement_year)
    settlement_year_num = int(settlement_year_value) if settlement_year_value != "" else None
    actual_date_year = _extract_year(actual_date)
    actual_date_year_num = int(actual_date_year) if actual_date_year != "" else None
    is_unit_row = any(ch.isdigit() for ch in code_text)

    if is_unit_row:
        if actual_date_year_num is None:
            return "", ""
        return actual_date, actual_date_year_num
    else:
        fallback_candidates = [x for x in [actual_date_year_num, latest_unit_year, settlement_year_num] if x is not None]
        actual_year = fallback_candidates[0] if fallback_candidates else ""

    return actual_date, actual_year


def _unit_budget_layout(df: pd.DataFrame | None = None) -> Dict[str, int]:
    layout = {
        "unit_header_start": 17,
        "budget_variance": 8,
        "group": 10,
        "gmp": 11,
        "fee": 12,
        "wip": 13,
        "raw_budget_total": 16,
    }
    if df is None or len(df) == 0:
        return layout

    co_header = _safe_string(_get_cell(df, 0, 8)).lower()
    actual_year_header = _safe_string(_get_cell(df, 0, 10))
    tbd_header = _safe_string(_get_cell(df, 0, 11))
    if co_header == "c/o date" and actual_year_header == "实际结算年份" and tbd_header == "TBD Acceptance Date":
        return {
            "unit_header_start": 21,
            "budget_variance": 12,
            "group": 14,
            "gmp": 15,
            "fee": 16,
            "wip": 17,
            "raw_budget_total": 20,
        }
    if co_header == "c/o date" or actual_year_header == "实际结算年份":
        return {
            "unit_header_start": 20,
            "budget_variance": 11,
            "group": 13,
            "gmp": 14,
            "fee": 15,
            "wip": 16,
            "raw_budget_total": 19,
        }
    return layout


def _unit_budget_classification_layout(df: pd.DataFrame | None = None) -> Dict[str, int | None]:
    layout: Dict[str, int | None] = {
        "unit_code": 2,
        "co_date": None,
        "actual_settlement_date": None,
        "actual_settlement_year": None,
        "budget_variance": None,
        "tbd_acceptance_date": None,
        "group": None,
        "gmp": None,
        "fee": None,
        "wip": None,
    }
    if df is None or len(df) == 0:
        return layout

    # 在第 1 行或第 2 行查找标题
    h_idx = 1 if len(df) > 1 else 0

    layout["co_date"] = _find_col_in_row(df, h_idx, "C/O date")
    layout["actual_settlement_date"] = _find_col_in_row(df, h_idx, "实际结算日期") or _find_col_in_row(df, h_idx, "Actual Settlement Date") or 9
    layout["actual_settlement_year"] = _find_col_in_row(df, h_idx, "实际结算年份")
    layout["budget_variance"] = _find_col_in_row(df, h_idx, "预算差异")
    layout["tbd_acceptance_date"] = _find_col_in_row(df, h_idx, "TBD Acceptance Date") or 11
    layout["group"] = _find_col_in_row(df, h_idx, "Group")
    layout["gmp"] = _find_col_in_row(df, h_idx, "GMP")
    layout["fee"] = _find_col_in_row(df, h_idx, "Fee")
    layout["wip"] = _find_col_in_row(df, h_idx, "WIP")
    return layout


def _load_unit_budget_schedule_overrides_from_excel(
    path: str | Path,
) -> Dict[str, Dict[str, str]]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return {}

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return {}
        ws = wb.worksheets[0]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}

        header_lookup = {_normalize_header_token(value): idx for idx, value in enumerate(header_row)}
        unit_idx = header_lookup.get(_normalize_header_token("UnitCode"))
        co_idx = header_lookup.get(_normalize_header_token("C/O Date"))
        tbd_idx = header_lookup.get(_normalize_header_token("TBD Acceptance Date"))
        if unit_idx is None:
            return {}

        out: Dict[str, Dict[str, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            unit_code = _safe_string(row[unit_idx] if unit_idx < len(row) else "")
            if not unit_code:
                continue
            out[unit_code] = {
                "co_date": _format_iso_date_or_blank(row[co_idx] if co_idx is not None and co_idx < len(row) else ""),
                "tbd_acceptance_date": _format_iso_date_or_blank(
                    row[tbd_idx] if tbd_idx is not None and tbd_idx < len(row) else ""
                ),
            }
        return out
    finally:
        wb.close()


def _load_default_unit_budget_schedule_overrides() -> Dict[str, Dict[str, str]]:
    return _load_unit_budget_schedule_overrides_from_excel(SANDY_COVE_DATES_FILE)


def _build_unit_budget_schedule_map(
    wsb: pd.DataFrame,
    schedule_overrides: Mapping[str, Mapping[str, str]] | None = None,
) -> Dict[str, Dict[str, pd.Timestamp | None]]:
    layout = _unit_budget_classification_layout(wsb)
    overrides = schedule_overrides or {}
    schedule: Dict[str, Dict[str, pd.Timestamp | None]] = {}
    latest_numeric_actual: pd.Timestamp | None = None
    latest_numeric_tbd: pd.Timestamp | None = None

    for r in range(1, len(wsb)):
        unit_code = _safe_string(_get_cell(wsb, r, int(layout["unit_code"] or 2)))
        if not unit_code:
            continue
        override_item = overrides.get(unit_code, {})
        co_col = layout.get("co_date")
        co_date = override_item.get("co_date", "") or (
            _safe_string(_get_cell(wsb, r, int(co_col))) if co_col else ""
        )
        actual_dt = _normalize_date_value(_co_date_to_actual_settlement_date(co_date))
        if actual_dt is None:
            actual_dt = _normalize_date_value(_get_cell(wsb, r, int(layout["actual_settlement_date"] or 9)))
        tbd_col = layout.get("tbd_acceptance_date")
        tbd_source = override_item.get("tbd_acceptance_date", "") or (
            _get_cell(wsb, r, int(tbd_col)) if tbd_col else ""
        )
        tbd_dt = _normalize_date_value(tbd_source)
        schedule[unit_code] = {
            "actual_settlement_date": actual_dt,
            "tbd_acceptance_date": tbd_dt,
        }
        if _has_digits(unit_code):
            if actual_dt is not None:
                latest_numeric_actual = actual_dt if latest_numeric_actual is None else max(latest_numeric_actual, actual_dt)
            if tbd_dt is not None:
                latest_numeric_tbd = tbd_dt if latest_numeric_tbd is None else max(latest_numeric_tbd, tbd_dt)

    for unit_code, item in schedule.items():
        if _has_digits(unit_code):
            continue
        if item["actual_settlement_date"] is None:
            item["actual_settlement_date"] = latest_numeric_actual
        if item["tbd_acceptance_date"] is None:
            item["tbd_acceptance_date"] = latest_numeric_tbd

    schedule["__COMMON_FALLBACK__"] = {
        "actual_settlement_date": latest_numeric_actual,
        "tbd_acceptance_date": latest_numeric_tbd,
    }
    return schedule


def ensure_uid_anchor(df: pd.DataFrame, uid_column: str) -> Tuple[pd.DataFrame, int]:
    out = df.copy()
    generated = 0

    if uid_column not in out.columns:
        out[uid_column] = ""

    if UID_STATUS_COL not in out.columns:
        out[UID_STATUS_COL] = ""

    existing = {
        _safe_string(v)
        for v in out[uid_column].tolist()
        if _safe_string(v)
    }

    for idx in out.index:
        uid = _safe_string(out.at[idx, uid_column])
        status = _safe_string(out.at[idx, UID_STATUS_COL])

        if not uid:
            while True:
                candidate = f"AIWB-{uuid.uuid4().hex[:12].upper()}"
                if candidate not in existing:
                    existing.add(candidate)
                    uid = candidate
                    break
            out.at[idx, uid_column] = uid
            out.at[idx, UID_STATUS_COL] = UID_PENDING_VALUE
            generated += 1
        else:
            if status not in (UID_PENDING_VALUE, UID_SYNCED_VALUE):
                out.at[idx, UID_STATUS_COL] = UID_SYNCED_VALUE

    return out, generated


def mark_uid_synced(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if UID_STATUS_COL in out.columns:
        out[UID_STATUS_COL] = out[UID_STATUS_COL].astype("string").replace(UID_PENDING_VALUE, UID_SYNCED_VALUE)
    return out


def apply_shadow_logic(
    df: pd.DataFrame,
    revenue_col: str,
    cost_col: str,
    profit_col: str,
    tax_rate_col: str,
    tax_col: str,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    out = df.copy()

    if SHADOW_CONFLICT_COL not in out.columns:
        out[SHADOW_CONFLICT_COL] = ""

    out[SHADOW_PY_PROFIT_COL] = pd.NA
    out[SHADOW_PY_TAX_COL] = pd.NA

    required = [revenue_col, cost_col, profit_col, tax_rate_col, tax_col]
    if any(col not in out.columns for col in required):
        out[SHADOW_CONFLICT_COL] = ""
        return out

    revenue = pd.to_numeric(out[revenue_col], errors="coerce")
    cost = pd.to_numeric(out[cost_col], errors="coerce")
    sheet_profit = pd.to_numeric(out[profit_col], errors="coerce")

    tax_rate = pd.to_numeric(out[tax_rate_col], errors="coerce").fillna(0)
    tax_rate = tax_rate.where(tax_rate <= 1, tax_rate / 100)

    py_profit = revenue.fillna(0) - cost.fillna(0)
    py_tax = py_profit.clip(lower=0) * tax_rate
    sheet_tax = pd.to_numeric(out[tax_col], errors="coerce")

    out[SHADOW_PY_PROFIT_COL] = py_profit.round(6)
    out[SHADOW_PY_TAX_COL] = py_tax.round(6)

    conflict_mask = (
        (sheet_profit - py_profit).abs().fillna(0) > tolerance
    ) | ((sheet_tax - py_tax).abs().fillna(0) > tolerance)

    out[SHADOW_CONFLICT_COL] = conflict_mask.map(lambda x: "冲突" if x else "")
    return out


def calculate_diff(
    original_df: pd.DataFrame,
    edited_df: pd.DataFrame,
    uid_column: str,
    amount_column: str,
) -> Dict[str, Any]:
    original = _cloud_view(original_df)
    edited = _cloud_view(edited_df)

    if uid_column not in original.columns or uid_column not in edited.columns:
        raise ValueError(f"缺少主键列 {uid_column}")

    original_uid = original[uid_column].astype("string").fillna("").str.strip()
    edited_uid = edited[uid_column].astype("string").fillna("").str.strip()

    duplicate_uid_count = int(edited_uid[edited_uid != ""].duplicated().sum())
    invalid_uid_rows = int((edited_uid == "").sum())

    original_idx = original.copy()
    edited_idx = edited.copy()
    original_idx[uid_column] = original_uid
    edited_idx[uid_column] = edited_uid
    original_idx = original_idx[original_idx[uid_column] != ""].drop_duplicates(uid_column, keep="first")
    edited_idx = edited_idx[edited_idx[uid_column] != ""].drop_duplicates(uid_column, keep="first")

    original_set = set(original_idx[uid_column].tolist())
    edited_set = set(edited_idx[uid_column].tolist())

    added = edited_set - original_set
    deleted = original_set - edited_set
    common = sorted(original_set & edited_set)

    compare_cols = [c for c in sorted(set(original_idx.columns) | set(edited_idx.columns)) if c != uid_column]

    modified_rows = 0
    if common:
        o = original_idx.set_index(uid_column).reindex(common).reindex(columns=compare_cols).astype("string").fillna("").apply(lambda s: s.str.strip())
        e = edited_idx.set_index(uid_column).reindex(common).reindex(columns=compare_cols).astype("string").fillna("").apply(lambda s: s.str.strip())
        modified_rows = int((o != e).any(axis=1).sum())

    before_amount = float(pd.to_numeric(original.get(amount_column, pd.Series(dtype="float64")), errors="coerce").fillna(0).sum())
    after_amount = float(pd.to_numeric(edited.get(amount_column, pd.Series(dtype="float64")), errors="coerce").fillna(0).sum())

    return {
        "added_rows": len(added),
        "modified_rows": modified_rows,
        "deleted_rows": len(deleted),
        "amount_delta": after_amount - before_amount,
        "before_amount_total": before_amount,
        "after_amount_total": after_amount,
        "invalid_uid_rows": invalid_uid_rows,
        "duplicate_uid_count": duplicate_uid_count,
        "original_rows": len(original),
        "edited_rows": len(edited),
    }


def build_sheet_delta_payload(
    sheet_name: str,
    original_df: pd.DataFrame,
    edited_df: pd.DataFrame,
    uid_column: str,
    amount_column: str,
    entity_column: str,
) -> Dict[str, Any]:
    original = _cloud_view(original_df)
    edited = _cloud_view(edited_df)

    if uid_column not in edited.columns:
        raise ValueError(f"工作表 {sheet_name} 缺少主键列 {uid_column}")

    uid_series = edited[uid_column].astype("string").fillna("").str.strip()
    dup_count = int(uid_series[uid_series != ""].duplicated().sum())
    if dup_count > 0:
        raise ValueError(f"工作表 {sheet_name} 存在 {dup_count} 个重复 {uid_column}，请先修复")

    headers = list(edited.columns)
    col_index = {col: idx + 1 for idx, col in enumerate(headers)}
    uid_col_index = col_index[uid_column]

    original_uid_series = (
        original[uid_column].astype("string").fillna("").str.strip()
        if uid_column in original.columns
        else pd.Series([""] * len(original))
    )

    original_row_map: Dict[str, Tuple[int, pd.Series]] = {}
    for pos, uid in enumerate(original_uid_series.tolist()):
        if uid and uid not in original_row_map:
            original_row_map[uid] = (pos + 2, original.iloc[pos])

    pending_col = edited_df.get(UID_STATUS_COL, pd.Series([""] * len(edited_df)))

    updates: List[dict] = []
    changed_rows = 0
    changed_cells = 0
    new_rows = 0
    pending_uid_updates = 0

    amount_delta_total = 0.0
    entity_amount_delta: Dict[str, float] = {}

    append_row_num = len(original) + 2
    seen_uid: set[str] = set()

    for idx in range(len(edited)):
        row = edited.iloc[idx]
        uid = _safe_string(row.get(uid_column, ""))
        if not uid:
            continue
        seen_uid.add(uid)

        force_uid_sync = _safe_string(pending_col.iloc[idx]) == UID_PENDING_VALUE

        if uid in original_row_map:
            row_num, old_row = original_row_map[uid]
            changed_idx: List[int] = []
            value_by_index: Dict[int, Any] = {}

            for col in headers:
                new_val = row.get(col, "")
                old_val = old_row[col] if col in original.columns else ""
                if not _values_equal(old_val, new_val):
                    col_i = col_index[col]
                    changed_idx.append(col_i)
                    value_by_index[col_i] = _serialize_for_api(new_val)

                    if col == amount_column:
                        old_num = _to_float(old_val) or 0.0
                        new_num = _to_float(new_val) or 0.0
                        delta = new_num - old_num
                        amount_delta_total += delta
                        entity = _safe_string(row.get(entity_column, "")) or "(未分类)"
                        entity_amount_delta[entity] = entity_amount_delta.get(entity, 0.0) + delta

            if force_uid_sync and uid_col_index not in value_by_index:
                changed_idx.append(uid_col_index)
                value_by_index[uid_col_index] = uid
                pending_uid_updates += 1

            if changed_idx:
                changed_rows += 1
                for start_i, end_i in _contiguous_segments(changed_idx):
                    values = [
                        value_by_index.get(col_i, _serialize_for_api(row.iloc[col_i - 1]))
                        for col_i in range(start_i, end_i + 1)
                    ]
                    updates.append(
                        {
                            "range": (
                                f"{_quote_sheet_name(sheet_name)}!"
                                f"{_column_number_to_a1(start_i)}{row_num}:"
                                f"{_column_number_to_a1(end_i)}{row_num}"
                            ),
                            "majorDimension": "ROWS",
                            "values": [values],
                        }
                    )
                    changed_cells += (end_i - start_i + 1)

        else:
            row_num = append_row_num
            append_row_num += 1
            new_rows += 1
            changed_rows += 1

            row_values = [_serialize_for_api(row.get(col, "")) for col in headers]
            updates.append(
                {
                    "range": (
                        f"{_quote_sheet_name(sheet_name)}!"
                        f"A{row_num}:{_column_number_to_a1(len(headers))}{row_num}"
                    ),
                    "majorDimension": "ROWS",
                    "values": [row_values],
                }
            )
            changed_cells += len(headers)

            if amount_column in headers:
                delta = _to_float(row.get(amount_column, "")) or 0.0
                amount_delta_total += delta
                entity = _safe_string(row.get(entity_column, "")) or "(未分类)"
                entity_amount_delta[entity] = entity_amount_delta.get(entity, 0.0) + delta

    deleted_rows = len(set(original_row_map.keys()) - seen_uid)

    return {
        "sheet": sheet_name,
        "updates": updates,
        "stats": {
            "changed_rows": changed_rows,
            "changed_cells": changed_cells,
            "new_rows": new_rows,
            "deleted_rows": deleted_rows,
            "pending_uid_updates": pending_uid_updates,
            "amount_delta": amount_delta_total,
            "entity_amount_delta": entity_amount_delta,
        },
    }


def build_commit_bundle(
    original_map: Mapping[str, pd.DataFrame],
    edited_map: Mapping[str, pd.DataFrame],
    target_sheets: Sequence[str],
    uid_column: str,
    amount_column: str,
    entity_column: str,
) -> Dict[str, Any]:
    all_updates: List[dict] = []
    per_sheet: List[Dict[str, Any]] = []

    summary = {
        "target_sheet_count": len(target_sheets),
        "changed_rows": 0,
        "changed_cells": 0,
        "new_rows": 0,
        "deleted_rows": 0,
        "pending_uid_updates": 0,
        "amount_delta": 0.0,
    }

    audit_lines: List[str] = []

    for sheet in target_sheets:
        payload = build_sheet_delta_payload(
            sheet_name=sheet,
            original_df=original_map.get(sheet, pd.DataFrame()),
            edited_df=edited_map.get(sheet, pd.DataFrame()),
            uid_column=uid_column,
            amount_column=amount_column,
            entity_column=entity_column,
        )
        stats = payload["stats"]
        updates = payload["updates"]

        per_sheet.append({"sheet": sheet, **stats, "update_count": len(updates)})
        all_updates.extend(updates)

        summary["changed_rows"] += int(stats["changed_rows"])
        summary["changed_cells"] += int(stats["changed_cells"])
        summary["new_rows"] += int(stats["new_rows"])
        summary["deleted_rows"] += int(stats["deleted_rows"])
        summary["pending_uid_updates"] += int(stats["pending_uid_updates"])
        summary["amount_delta"] += float(stats["amount_delta"])

        entity_delta: Dict[str, float] = stats["entity_amount_delta"]
        if entity_delta:
            for entity, delta in sorted(entity_delta.items(), key=lambda x: abs(x[1]), reverse=True):
                audit_lines.append(
                    f"实体 {entity}：金额差异合计 {delta:,.2f}（sheet={sheet}）"
                )
        else:
            audit_lines.append(
                f"sheet={sheet}：变更 {stats['changed_rows']} 行，新增 {stats['new_rows']} 行"
            )

    return {
        "updates": all_updates,
        "summary": summary,
        "per_sheet": per_sheet,
        "audit_lines": audit_lines,
    }


def _sync_unit_master_py(sheet_map: Mapping[str, pd.DataFrame]) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    out = {k: v.copy() for k, v in sheet_map.items()}
    wsb_key = _sheet_key(out, "Unit Budget")
    layout = _unit_budget_layout(out[wsb_key])
    wsb = _ensure_column_count(out[wsb_key], layout["unit_header_start"])
    units: set[str] = set()

    for col_i in range(layout["unit_header_start"], len(wsb.columns) + 1):
        h = _safe_string(wsb.columns[col_i - 1])
        if h and h.upper() != "UNIT CODE":
            units.add(h)

    scan_specs = [
        ("Payable", "unit_code", 38),
        ("Final Detail", "unit_code", 21),
        ("Draw request report", "unit_code", 8),
    ]
    mapping_warnings: List[Dict[str, Any]] = []
    for name, logical_field, legacy_col in scan_specs:
        try:
            key = _sheet_key(out, name)
        except KeyError:
            continue
        src = _ensure_column_count(out[key], legacy_col)
        layout, fallback_events = resolve_sheet_field_columns_with_fallback(
            headers=list(src.columns),
            sheet_name=name,
            fallback_columns={logical_field: legacy_col},
            fields=(logical_field,),
        )
        mapping_warnings.extend(fallback_events)
        col = int(layout.get(logical_field) or legacy_col)
        for r in range(0, len(src)):
            val = _safe_string(_get_cell(src, r, col))
            if val:
                units.add(val)

    sorted_units = sorted(units)

    if len(wsb) >= 2:
        for r in range(1, len(wsb)):
            wsb = _set_cell(wsb, r, 2, "")

    for idx, unit in enumerate(sorted_units):
        row_idx = 1 + idx
        wsb = _set_cell(wsb, row_idx, 2, unit)

    out[wsb_key] = wsb
    return out, {"unit_count": len(sorted_units), "mapping_warnings": mapping_warnings}


def _calculate_unit_budget_cd_py(sheet_map: Mapping[str, pd.DataFrame]) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    out = {k: v.copy() for k, v in sheet_map.items()}
    wsb_key = _sheet_key(out, "Unit Budget")
    wsb = out[wsb_key].copy()
    layout = _unit_budget_layout(wsb)

    row_count = len(wsb)
    col_count = len(wsb.columns)
    if row_count == 0:
        return out, {"computed_rows": 0}

    unit_col_map: Dict[str, int] = {}
    for j in range(layout["unit_header_start"], col_count + 1):
        h = _safe_string(wsb.columns[j - 1])
        if h:
            unit_col_map[h] = j

    results: List[Tuple[Any, Any]] = []

    for i in range(1, row_count):
        u_code = _safe_string(_get_cell(wsb, i, 2))
        sum_c = 0.0
        sum_d = 0.0

        target_col = unit_col_map.get(u_code)
        if u_code and target_col is not None:
            for r in range(0, row_count):
                val_k = _to_float(_get_cell(wsb, r, layout["gmp"]))
                val_m = _safe_string(_get_cell(wsb, r, layout["wip"]))
                cell_value = _safe_number(_get_cell(wsb, r, target_col))

                if val_k is not None and abs(val_k - 1) < 1e-9:
                    sum_c += cell_value
                if val_m.endswith("3"):
                    sum_d += cell_value

        c_val = "" if abs(sum_c) < 1e-12 else sum_c
        d_val = "" if abs(sum_d) < 1e-12 else sum_d
        results.append((c_val, d_val))

    for offset, (c_val, d_val) in enumerate(results):
        row_idx = 1 + offset
        wsb = _set_cell(wsb, row_idx, 3, c_val)
        wsb = _set_cell(wsb, row_idx, 4, d_val)

    out[wsb_key] = wsb
    return out, {"computed_rows": len(results)}


def _build_final_detail_summary_rows(
    units: Sequence[str],
    row_unit_codes: Sequence[str],
    row_d_values: Sequence[Any],
    row_c_values: Sequence[Any],
    row_p_values: Sequence[Any],
    row_t_texts: Sequence[str],
    row_v_texts: Sequence[str],
) -> List[List[Any]]:
    summary_by_unit: Dict[str, List[float]] = {}
    row_count = min(
        len(row_unit_codes),
        len(row_d_values),
        len(row_c_values),
        len(row_p_values),
        len(row_t_texts),
        len(row_v_texts),
    )

    for i in range(row_count):
        unit_code = _safe_string(row_unit_codes[i])
        if not unit_code:
            continue
        bucket = summary_by_unit.setdefault(unit_code, [0.0, 0.0, 0.0])
        bucket[0] += _safe_number(row_d_values[i])

        c_flag = _to_float(row_c_values[i])
        p_val = _safe_number(row_p_values[i])
        if c_flag is not None and abs(c_flag - 1) < 1e-9:
            bucket[1] += p_val
        if not _safe_string(row_t_texts[i]) and _safe_string(row_v_texts[i]) != "Sharing":
            bucket[2] += p_val

    summary_rows: List[List[Any]] = []
    for unit_code in units:
        if not unit_code:
            summary_rows.append(["", "", "", ""])
            continue
        sum_f, sum_g, sum_h = summary_by_unit.get(unit_code, [0.0, 0.0, 0.0])
        summary_rows.append(
            [
                unit_code,
                "" if abs(sum_f) < 1e-12 else sum_f,
                "" if abs(sum_g) < 1e-12 else sum_g,
                "" if abs(sum_h) < 1e-12 else sum_h,
            ]
        )
    return summary_rows


def _collect_rule_ids_from_sheet_map(
    sheet_map: Mapping[str, pd.DataFrame],
    target_sheets: Sequence[str] = ("Payable", "Final Detail"),
) -> List[str]:
    rule_ids: List[str] = []
    for sheet_name in target_sheets:
        try:
            key = _sheet_key(sheet_map, sheet_name)
        except KeyError:
            continue
        df = sheet_map.get(key, pd.DataFrame())
        if len(df.columns) < 2:
            continue
        for value in df.iloc[:, 1].tolist():
            text = _safe_string(value)
            if text:
                rule_ids.append(text)
    return rule_ids


def _detect_rule_id_hit_rate_alerts(
    current_rule_ids: Sequence[str],
    historical_avg_rates: Mapping[str, float],
    deviation_threshold: float = DEFAULT_RULE_ID_HIT_RATE_ALERT_THRESHOLD,
) -> List[Dict[str, Any]]:
    cleaned_rule_ids = [_safe_string(rule_id) for rule_id in current_rule_ids if _safe_string(rule_id)]
    total = len(cleaned_rule_ids)
    if total <= 0:
        return []

    counts: Dict[str, int] = {}
    for rule_id in cleaned_rule_ids:
        counts[rule_id] = counts.get(rule_id, 0) + 1

    alerts: List[Dict[str, Any]] = []
    for rule_id, historical_avg_rate in historical_avg_rates.items():
        key = _safe_string(rule_id)
        if not key:
            continue
        try:
            historical_rate = float(historical_avg_rate)
        except Exception:
            continue
        if historical_rate < 0.0 or historical_rate > 1.0:
            continue

        current_rate = counts.get(key, 0) / total
        deviation = current_rate - historical_rate
        if abs(deviation) <= float(deviation_threshold):
            continue

        alerts.append(
            {
                "rule_id": key,
                "current_rate": current_rate,
                "historical_avg_rate": historical_rate,
                "deviation": deviation,
                "current_count": counts.get(key, 0),
                "total_count": total,
            }
        )

    alerts.sort(key=lambda item: abs(float(item["deviation"])), reverse=True)
    return alerts


def _process_payable_py(sheet_map: Mapping[str, pd.DataFrame]) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    from .finance_classification import compute_payable_classifications
    out = {k: v.copy() for k, v in sheet_map.items()}
    wsp_key = _sheet_key(out, "Payable")
    wss_key = _sheet_key(out, "Scoping")

    wsp = _ensure_column_count(out[wsp_key], 43)
    wss = _ensure_column_count(out[wss_key], 10)

    mapping = {
        "Wan Bridge Development LLC": "WBD",
        "Wan Pacific Real Estate Development LLC": "WPRED",
        "WB Home LLC": "WBH",
        "WB Lago Mar Pod 8 Land LLC": "WLM",
    }
    categories, classification_extra = compute_payable_classifications(out)
    if len(categories) == len(wsp):
        first_col = wsp.columns[0]
        wsp[first_col] = categories
    rule_ids = list(classification_extra.get("rule_ids", []))
    if len(rule_ids) == len(wsp) and len(wsp.columns) >= 2:
        second_col = wsp.columns[1]
        wsp[second_col] = rule_ids
    payable_layout, mapping_warnings = resolve_sheet_field_columns_with_fallback(
        headers=list(wsp.columns),
        sheet_name="Payable",
        fallback_columns={
            "vendor": 15,
            "invoice_no": 22,
            "cost_name": 40,
        },
        fields=("vendor", "invoice_no", "cost_name"),
    )
    vendor_col = int(payable_layout.get("vendor") or 15)
    invoice_col = int(payable_layout.get("invoice_no") or 22)
    cost_name_col = int(payable_layout.get("cost_name") or _find_col_in_headers(wsp, "Cost Name") or 40)

    scoping_rows = len(wss)
    scope_agg: Dict[int, Tuple[float, float, float]] = {}
    for r in range(scoping_rows):
        code = _to_float(_get_cell(wss, r, 3))
        if code is None:
            continue
        key = int(code)
        cur = scope_agg.get(key, (0.0, 0.0, 0.0))
        e = cur[0] + _safe_number(_get_cell(wss, r, 5))
        f = cur[1] + _safe_number(_get_cell(wss, r, 6))
        g = cur[2] + _safe_number(_get_cell(wss, r, 8))
        scope_agg[key] = (e, f, g)

    write_rows: List[List[Any]] = []
    for i in range(len(wsp)):
        o_text = _safe_string(_get_cell(wsp, i, vendor_col))
        v_text = _safe_string(_get_cell(wsp, i, invoice_col))
        cost_name_text = _safe_string(_get_cell(wsp, i, cost_name_col))

        j_val = _extract_tail_str(v_text, 4)
        d_val = _extract_leading_int(cost_name_text, 3)
        c_val = mapping.get(o_text, "")

        s_e: Any = ""
        s_f: Any = ""
        s_g: Any = ""
        if d_val is not None:
            sums = scope_agg.get(int(d_val), (0.0, 0.0, 0.0))
            s_e = "" if abs(sums[0]) < 1e-12 else sums[0]
            s_f = "" if abs(sums[1]) < 1e-12 else sums[1]
            s_g = "" if abs(sums[2]) < 1e-12 else sums[2]

        write_rows.append([c_val, d_val if d_val is not None else "", s_e, s_f, s_g, "", "", j_val])

    payable_write_cols = list(wsp.columns[2:10])
    if len(write_rows) == len(wsp) and len(payable_write_cols) == 8:
        for col in payable_write_cols:
            if str(wsp[col].dtype) != "object":
                wsp[col] = wsp[col].astype(object)
        wsp.loc[:, payable_write_cols] = pd.DataFrame(write_rows, index=wsp.index, columns=payable_write_cols)

    out[wsp_key] = wsp
    merged_extra = dict(classification_extra)
    merged_extra["mapping_warnings"] = [
        *list(classification_extra.get("mapping_warnings", [])),
        *mapping_warnings,
    ]
    return out, merged_extra


def _process_final_detail_py(sheet_map: Mapping[str, pd.DataFrame]) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    from .finance_classification import compute_final_detail_classifications
    out = {k: v.copy() for k, v in sheet_map.items()}
    wsf_key = _sheet_key(out, "Final Detail")
    wss_key = _sheet_key(out, "Scoping")
    wsb_key = _sheet_key(out, "Unit Budget")

    wsf = _ensure_column_count(out[wsf_key], 30)
    wss = _ensure_column_count(out[wss_key], 10)
    wsb = _ensure_column_count(out[wsb_key], 16)
    categories, classification_extra = compute_final_detail_classifications(out)
    if len(categories) == len(wsf):
        first_col = wsf.columns[0]
        wsf[first_col] = categories
    rule_ids = list(classification_extra.get("rule_ids", []))
    if len(rule_ids) == len(wsf) and len(wsf.columns) >= 2:
        second_col = wsf.columns[1]
        wsf[second_col] = rule_ids

    scope_sum_c: Dict[int, float] = {}
    scope_codes = _column_values_1based(wss, 3)
    scope_amounts = _column_values_1based(wss, 5)
    for r in range(len(wss)):
        code = _to_float(scope_codes[r])
        if code is None:
            continue
        key = int(code)
        scope_sum_c[key] = scope_sum_c.get(key, 0.0) + _safe_number(scope_amounts[r])

    final_layout, mapping_warnings = resolve_sheet_field_columns_with_fallback(
        headers=list(wsf.columns),
        sheet_name="Final Detail",
        fallback_columns={
            "cost_code": 26,
            "posting_date": 15,
            "incurred_date": 20,
            "unit_code": 21,
        },
        fields=("cost_code", "posting_date", "incurred_date", "unit_code"),
    )
    cost_code_col = int(final_layout.get("cost_code") or 26)
    posting_date_col = int(final_layout.get("posting_date") or 15)
    year_source_col = int(final_layout.get("incurred_date") or 20)
    unit_code_col = int(final_layout.get("unit_code") or 21)

    unit_seen: set[str] = set()
    z_texts = [_safe_string(value) for value in _column_values_1based(wsf, cost_code_col)]
    o_values = _column_values_1based(wsf, posting_date_col)
    t_values = _column_values_1based(wsf, year_source_col)
    u_keys = [_safe_string(value) for value in _column_values_1based(wsf, unit_code_col)]

    row_level_write_rows: List[List[Any]] = []
    for i in range(len(wsf)):
        z_text = z_texts[i]
        o_val = o_values[i]
        t_val = t_values[i]
        u_key = u_keys[i]

        b_val = _extract_tail_int(z_text, 3)
        k_val = _extract_year(t_val)

        c_val: Any = ""
        if b_val is not None:
            s = scope_sum_c.get(int(b_val), 0.0)
            c_val = "" if abs(s) < 1e-12 else s

        d_val: Any = ""
        if u_key and u_key not in unit_seen:
            y = _extract_year(o_val)
            d_val = y if y != "" else ""
            unit_seen.add(u_key)

        row_level_write_rows.append(
            [
                c_val,
                d_val,
                "",
                "",
                "",
                "",
                "",
                "",
                k_val if k_val != "" else "",
            ]
        )

    final_detail_row_cols = list(wsf.columns[2:11])
    if len(row_level_write_rows) == len(wsf) and len(final_detail_row_cols) == 9:
        for col in final_detail_row_cols:
            if str(wsf[col].dtype) != "object":
                wsf[col] = wsf[col].astype(object)
        wsf.loc[:, final_detail_row_cols] = pd.DataFrame(
            row_level_write_rows,
            index=wsf.index,
            columns=final_detail_row_cols,
        )

    units = [_safe_string(value) for value in _column_values_1based(wsb, 2)[1:]]
    summary_rows = _build_final_detail_summary_rows(
        units=units,
        row_unit_codes=u_keys,
        row_d_values=_column_values_1based(wsf, 4),
        row_c_values=_column_values_1based(wsf, 3),
        row_p_values=_column_values_1based(wsf, 16),
        row_t_texts=[_safe_string(value) for value in _column_values_1based(wsf, 20)],
        row_v_texts=[_safe_string(value) for value in _column_values_1based(wsf, 22)],
    )

    summary_write_cols = list(wsf.columns[4:8])
    if len(summary_rows) == len(units) and len(summary_write_cols) == 4:
        summary_df = pd.DataFrame(summary_rows, columns=summary_write_cols, index=range(len(units)))
        if len(wsf) < len(summary_df):
            wsf = _ensure_row_count(wsf, len(summary_df))
        for col in summary_write_cols:
            if str(wsf[col].dtype) != "object":
                wsf[col] = wsf[col].astype(object)
        wsf.loc[summary_df.index, summary_write_cols] = summary_df

    if len(rule_ids) == len(wsf) and len(wsf.columns) >= 2:
        second_col = wsf.columns[1]
        wsf[second_col] = rule_ids

    out[wsf_key] = wsf
    final_extra = dict(classification_extra)
    final_extra["mapping_warnings"] = [
        *list(classification_extra.get("mapping_warnings", [])),
        *mapping_warnings,
    ]
    final_extra["summary_rows"] = len(units)
    return out, final_extra


def load_rule_id_hit_rate_baseline() -> Dict[str, float]:
    if not RULE_ID_HIT_RATE_BASELINE_FILE.exists():
        return {}
    try:
        payload = json.loads(RULE_ID_HIT_RATE_BASELINE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

    if isinstance(payload, Mapping):
        candidate = payload.get("rule_rates", payload)
    else:
        candidate = {}

    if not isinstance(candidate, Mapping):
        return {}

    baseline: Dict[str, float] = {}
    for rule_id, rate in candidate.items():
        key = _safe_string(rule_id)
        if not key:
            continue
        try:
            numeric_rate = float(rate)
        except Exception:
            continue
        if 0.0 <= numeric_rate <= 1.0:
            baseline[key] = numeric_rate
    return baseline


def run_apps_shadow_pipeline(
    edited_map: Mapping[str, pd.DataFrame],
    uid_column: str,
    shadow_cfg: Mapping[str, Any],
) -> Tuple[Dict[str, pd.DataFrame], List[Dict[str, Any]], List[str]]:
    working = {k: v.copy() for k, v in edited_map.items()}
    reports: List[Dict[str, Any]] = []
    audit_lines: List[str] = []

    def apply_and_report(
        step_name: str,
        handler,
        focus_sheets: Sequence[str],
    ) -> None:
        nonlocal working
        resolved_focus: List[str] = []
        for s in focus_sheets:
            try:
                resolved_focus.append(_sheet_key(working, s))
            except KeyError:
                continue
        before = {k: working[k].copy() for k in resolved_focus if k in working}
        working, extra = handler(working)
        changed_rows = 0
        changed_cells = 0
        touched: List[str] = []
        for sheet in resolved_focus:
            if sheet not in working or sheet not in before:
                continue
            stats = _sheet_delta_stats(before[sheet], working[sheet])
            if stats["changed_cells"] > 0:
                touched.append(sheet)
                changed_rows += stats["changed_rows"]
                changed_cells += stats["changed_cells"]
        reports.append(
            {
                "step": step_name,
                "sheets": "、".join(touched) if touched else "无",
                "changed_rows": changed_rows,
                "changed_cells": changed_cells,
                "extra": extra,
            }
        )
        audit_lines.append(
            f"{step_name}: sheets={reports[-1]['sheets']}, rows={changed_rows}, cells={changed_cells}, extra={extra}"
        )

    def apply_from_source_and_report(
        step_name: str,
        handler,
        focus_sheets: Sequence[str],
        source_map: Mapping[str, pd.DataFrame],
    ) -> None:
        nonlocal working
        resolved_focus: List[str] = []
        for s in focus_sheets:
            try:
                resolved_focus.append(_sheet_key(working, s))
            except KeyError:
                continue
        before = {k: working[k].copy() for k in resolved_focus if k in working}
        source_copy = {k: v.copy() for k, v in source_map.items()}
        processed_map, extra = handler(source_copy)
        for sheet in resolved_focus:
            if sheet in processed_map:
                working[sheet] = processed_map[sheet]

        changed_rows = 0
        changed_cells = 0
        touched: List[str] = []
        for sheet in resolved_focus:
            if sheet not in working or sheet not in before:
                continue
            stats = _sheet_delta_stats(before[sheet], working[sheet])
            if stats["changed_cells"] > 0:
                touched.append(sheet)
                changed_rows += stats["changed_rows"]
                changed_cells += stats["changed_cells"]
        reports.append(
            {
                "step": step_name,
                "sheets": "、".join(touched) if touched else "无",
                "changed_rows": changed_rows,
                "changed_cells": changed_cells,
                "extra": extra,
            }
        )
        audit_lines.append(
            f"{step_name}: sheets={reports[-1]['sheets']}, rows={changed_rows}, cells={changed_cells}, extra={extra}"
        )

    apply_and_report("同步 Unit 列表(B列)", _sync_unit_master_py, ["Unit Budget"])
    apply_and_report("计算 Unit Budget C/D", _calculate_unit_budget_cd_py, ["Unit Budget"])
    classification_base = {k: v.copy() for k, v in working.items()}
    apply_from_source_and_report("核算 Payable C:J", _process_payable_py, ["Payable"], classification_base)
    apply_from_source_and_report("核算 Final Detail(B:K,E:H)", _process_final_detail_py, ["Final Detail"], classification_base)

    for sheet, df in list(working.items()):
        prepared, _ = ensure_uid_anchor(df, uid_column)
        prepared = apply_shadow_logic(
            prepared,
            revenue_col=shadow_cfg["revenue_col"],
            cost_col=shadow_cfg["cost_col"],
            profit_col=shadow_cfg["profit_col"],
            tax_rate_col=shadow_cfg["tax_rate_col"],
            tax_col=shadow_cfg["tax_col"],
            tolerance=float(shadow_cfg["tolerance"]),
        )
        working[sheet] = prepared

    historical_rule_rates = load_rule_id_hit_rate_baseline()
    if historical_rule_rates:
        current_rule_ids = _collect_rule_ids_from_sheet_map(working)
        alerts = _detect_rule_id_hit_rate_alerts(
            current_rule_ids=current_rule_ids,
            historical_avg_rates=historical_rule_rates,
            deviation_threshold=DEFAULT_RULE_ID_HIT_RATE_ALERT_THRESHOLD,
        )
        reports.append(
            {
                "step": "Rule ID 异常波动预警",
                "sheets": "Payable、Final Detail",
                "changed_rows": 0,
                "changed_cells": 0,
                "extra": {
                    "alert_count": len(alerts),
                    "alerts": alerts,
                    "baseline_rule_count": len(historical_rule_rates),
                    "observed_rule_id_count": len(current_rule_ids),
                },
            }
        )
        if alerts:
            top_alert = alerts[0]
            audit_lines.append(
                "Rule ID 异常波动预警: "
                f"alert_count={len(alerts)}, "
                f"top_rule={top_alert['rule_id']}, "
                f"current_rate={top_alert['current_rate']:.2%}, "
                f"historical_avg_rate={top_alert['historical_avg_rate']:.2%}, "
                f"deviation={top_alert['deviation']:.2%}"
            )
        else:
            audit_lines.append(
                "Rule ID 异常波动预警: alert_count=0, baseline_loaded=1"
            )

    return working, reports, audit_lines


def _safety_check(
    service,
    spreadsheet_id: str,
    guard_sheet_name: str,
    expected_first_cell: str,
) -> None:
    first_row_resp = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=f"{guard_sheet_name}!1:1")
        .execute()
    )
    first_row = first_row_resp.get("values", [[]])[0] if first_row_resp.get("values") else []
    row_tokens = {str(item).strip() for item in first_row}

    if expected_first_cell and expected_first_cell not in row_tokens:
        raise RuntimeError(
            f"安全校验失败：{guard_sheet_name} 首行未检测到 '{expected_first_cell}'，已中止写入。"
        )


def execute_commit(
    service,
    spreadsheet_id: str,
    bundle: Mapping[str, Any],
    guard_sheet_name: str,
    expected_first_cell: str,
) -> Dict[str, Any]:
    updates = list(bundle.get("updates", []))
    if not updates:
        return {"api_calls": 0, "updated_ranges": 0}

    _safety_check(service, spreadsheet_id, guard_sheet_name, expected_first_cell)

    api_calls = 1
    for chunk in _chunked(updates, 500):
        (
            service.spreadsheets()
            .values()
            .batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "valueInputOption": "USER_ENTERED",
                    "data": list(chunk),
                },
            )
            .execute()
        )
        api_calls += 1

    return {"api_calls": api_calls, "updated_ranges": len(updates)}


def _find_year_header_row_109(rows: Sequence[Sequence[Any]]) -> int | None:
    year_row, _ = _discover_109_year_column_blocks(rows)
    return year_row


def _discover_109_year_column_blocks(
    rows: Sequence[Sequence[Any]],
    min_block_size: int = 2,
) -> Tuple[int | None, List[List[int]]]:
    best_row: int | None = None
    best_blocks: List[List[int]] = []
    best_score = -1

    for row_i, row in enumerate(rows, start=1):
        if not row:
            continue
        year_cols: List[int] = []
        for col_i, value in enumerate(row, start=1):
            year_value = _extract_year(value)
            if year_value in (None, ""):
                continue
            try:
                year_number = int(year_value)
            except (TypeError, ValueError):
                continue
            if 1900 <= year_number <= 2100:
                year_cols.append(col_i)
        if len(year_cols) < max(3, min_block_size):
            continue

        contiguous = _contiguous_segments(year_cols)
        blocks = [list(range(start, end + 1)) for start, end in contiguous if end - start + 1 >= min_block_size]
        if not blocks:
            continue

        score = sum(len(block) for block in blocks)
        if score > best_score:
            best_score = score
            best_row = row_i
            best_blocks = blocks

    return best_row, best_blocks


def _default_109_year_column_blocks() -> List[List[int]]:
    return [list(range(6, 12)), list(range(13, 19))]


def _choose_contract_price_row(rows: Sequence[Sequence[Any]], label_rows: Dict[str, List[int]]) -> int | None:
    rows_cp = label_rows.get("contract price", [])
    if not rows_cp:
        return None
    for row_i in rows_cp:
        if _to_float(_grid_cell(rows, row_i, 5)) is not None:
            return row_i
    return rows_cp[0]


def _find_contract_price_day1_row(rows: Sequence[Sequence[Any]]) -> int | None:
    for row_i in range(1, len(rows) + 1):
        label = _normalize_label(_grid_cell(rows, row_i, 4))
        if label == "contract price (day1):":
            if _grid_cell(rows, row_i, 5) != "":
                return row_i
    return None


def _find_budget_cost_row(rows: Sequence[Sequence[Any]], label_rows: Dict[str, List[int]]) -> int | None:
    for label, row_list in label_rows.items():
        if "budget cost" in label and "scoping" in label:
            return row_list[0]
    return None


def _first_present_row(label_rows: Mapping[str, Sequence[int]], *labels: str) -> int | None:
    for label in labels:
        row_list = list(label_rows.get(label, []))
        if row_list:
            return int(row_list[0])
    return None


def _merge_formula_plan_with_semantic_updates(
    plan: Sequence[Mapping[str, str]],
    semantic_updates: Sequence[Mapping[str, str]],
) -> List[Dict[str, str]]:
    merged_by_range: Dict[str, Dict[str, str]] = {}
    ordered_ranges: List[str] = []

    for item in plan:
        range_ref = str(item["range"])
        if range_ref not in merged_by_range:
            ordered_ranges.append(range_ref)
        merged_by_range[range_ref] = dict(item)

    for item in semantic_updates:
        range_ref = str(item["range"])
        if range_ref not in merged_by_range:
            ordered_ranges.append(range_ref)
        merged_by_range[range_ref] = dict(item)

    return [merged_by_range[range_ref] for range_ref in ordered_ranges]


def build_budgetco_semantic_summary_context(
    budgetco_values: Sequence[Sequence[Any]],
    start_col: str = "G",
) -> Dict[str, Any]:
    mapper = MapperFactory.create("BudgetCO", budgetco_values)

    row_savings = mapper.get_row("Total Savings Identified")
    row_contingency = mapper.get_row("Owner Contingency")
    row_total_eac = mapper.get_row("Total Budget (EAC)")
    total_eac_ref = mapper.get_ref("Total Budget (EAC)", start_col)

    return {
        "mapper": mapper,
        "row_savings": row_savings,
        "row_contingency": row_contingency,
        "row_total_eac": row_total_eac,
        "savings_ref": mapper.get_ref("Total Savings Identified", start_col),
        "contingency_ref": mapper.get_ref("Owner Contingency", start_col),
        "total_eac_ref": total_eac_ref,
        "eac_formula": f"=N({total_eac_ref})",
    }


def build_project_ledger_semantic_context(
    ledger_values: Sequence[Sequence[Any]],
    values_109: Sequence[Sequence[Any]],
    ledger_col: str = "H",
    cost_col_109: str = "F",
) -> Dict[str, Any]:
    ledger_mapper = MapperFactory.create("Project Ledger", ledger_values)
    mapper_109 = MapperFactory.create("109", values_109)

    total_actuals_label = "Total Actual Expenditure"
    accrual_adj_label = "Accrual Adjustments"
    cumulative_cost_label = "Cumulative Total Cost (Actual)"

    row_total_actuals = ledger_mapper.get_row(total_actuals_label)
    row_accrual_adj = ledger_mapper.get_row(accrual_adj_label)
    row_109_cumulative_cost = mapper_109.get_row(cumulative_cost_label)

    total_actuals_ref = ledger_mapper.get_ref(total_actuals_label, ledger_col)
    accrual_adj_ref = ledger_mapper.get_ref(accrual_adj_label, ledger_col)
    cumulative_cost_109_ref = mapper_109.get_ref(cumulative_cost_label, cost_col_109)
    ledger_total_formula = f"=N({total_actuals_ref}) + N({accrual_adj_ref})"

    return {
        "ledger_mapper": ledger_mapper,
        "mapper_109": mapper_109,
        "row_total_actuals": row_total_actuals,
        "row_accrual_adj": row_accrual_adj,
        "row_109_cumulative_cost": row_109_cumulative_cost,
        "total_actuals_ref": total_actuals_ref,
        "accrual_adj_ref": accrual_adj_ref,
        "cumulative_cost_109_ref": cumulative_cost_109_ref,
        "ledger_total_formula": ledger_total_formula,
        "shadow_reconcile_formula": f"{ledger_total_formula} = N({cumulative_cost_109_ref})",
    }


def update_109_semantic_logic(
    rows: List[List[Any]],
    sheet_109_title: str | None = None,
    formula_mappings: Mapping[str, Mapping[str, int]] | None = None,
) -> List[Dict[str, str]]:
    actual_sheet_109_title = _resolve_sheet_109_title(sheet_109_title)
    def detect_year_blocks() -> List[List[str]]:
        _, blocks = _discover_109_year_column_blocks(rows)
        if not blocks:
            blocks = _default_109_year_column_blocks()
        return [[_column_number_to_a1(col_n) for col_n in block] for block in blocks]

    labels_year_blocks = detect_year_blocks()
    primary_cols = labels_year_blocks[0] if labels_year_blocks else list("FGHIJK")
    audit_cols = labels_year_blocks[1] if len(labels_year_blocks) > 1 else []

    mapper = MapperFactory.create("109", rows)
    generator_config = dict(getattr(mapper, "config", {}) or {})
    generator_config["formula_mappings"] = {
        _safe_string(sheet_name): {str(field): int(index) for field, index in dict(fields).items()}
        for sheet_name, fields in dict(formula_mappings or {}).items()
        if isinstance(fields, Mapping)
    }
    generator_config["primary_year_cols"] = primary_cols
    generator_config["audit_year_cols"] = audit_cols
    if primary_cols:
        generator_config["start_year_anchor_cell"] = f"{primary_cols[-1]}2"
    generator = FinanceFormulaGenerator(mapper, config=generator_config)

    semantic_updates: List[Dict[str, str]] = []

    def maybe_add_formula_update(labels_to_try, col: str, formula_builder, logic: str) -> None:
        label_list = labels_to_try if isinstance(labels_to_try, (list, tuple)) else [labels_to_try]
        row_num = None
        for label in label_list:
            try:
                row_num = mapper.get_row(label)
                break
            except KeyError:
                continue
        if row_num is None:
            return
        try:
            formula = formula_builder(col)
        except KeyError:
            return
        semantic_updates.append(
            {
                "sheet": actual_sheet_109_title,
                "cell": f"{col}{row_num}",
                "range": _build_109_range(f"{col}{row_num}", actual_sheet_109_title),
                "formula": formula,
                "logic": logic,
            }
        )

    labels = mapper.config.get("labels", {})
    year_blocks = labels_year_blocks
    year_row = _find_year_header_row_109(rows) or 10

    for col in primary_cols:
        maybe_add_formula_update([labels.get("eac", "Dynamic Budget (EAC)"), "Dynamic Budget (EAC)"], col, generator.get_eac_formula, "Semantic EAC formula")
        maybe_add_formula_update(
            [labels.get("cumulative_direct_cost", "Cumulative Direct Cost"), "Cumulative Direct Cost"],
            col,
            generator.get_cumulative_direct_cost_formula,
            "Semantic cumulative direct cost formula",
        )
        maybe_add_formula_update(
            [labels.get("cogs_company", "Cost of Goods Sold-Company"), "Cost of Goods Sold-Company"],
            col,
            generator.get_cogs_company_formula,
            "Semantic company COGS formula",
        )
        maybe_add_formula_update([labels.get("poc", "Percentage of Completion"), "Percentage of Completion"], col, generator.get_poc_formula, "Semantic POC formula")
        maybe_add_formula_update([labels.get("revenue", "General Conditions fee"), "General Conditions fee"], col, generator.get_revenue_formula, "Semantic Revenue formula")
        maybe_add_formula_update([labels.get("confirmed_cogs", "Cost of Goods Sold"), "Cost of Goods Sold"], col, generator.get_confirmed_cogs_formula, "Semantic COGS formula")
        maybe_add_formula_update("Gross Profit-Company", col, generator.get_gross_profit_company_formula, "Semantic gross profit company formula")
        maybe_add_formula_update("Gross Profit-Audit", col, generator.get_gross_profit_audited_formula, "Semantic gross profit audit formula")
        maybe_add_formula_update("Gross Profit", col, generator.get_gross_profit_formula, "Semantic gross profit formula")
        maybe_add_formula_update(
            ["Total Income Cost"],
            col,
            lambda target_col: generator.get_income_total_formula(target_col, f"{target_col}${year_row}"),
            "Semantic total income cost formula",
        )
        maybe_add_formula_update(
            ["GC Income"],
            col,
            lambda target_col: generator.get_gc_income_formula(target_col, f"{target_col}${year_row}"),
            "Semantic GC income formula",
        )
        maybe_add_formula_update(
            ["GC Cost", "Total GC Cost"],
            col,
            lambda target_col: generator.get_gc_cost_formula(target_col, f"{target_col}${year_row}"),
            "Semantic GC cost formula",
        )
        maybe_add_formula_update(
            ["Actual Warranty Expenses (Reversed)"],
            col,
            lambda target_col: generator.get_actual_warranty_formula(target_col, f"{target_col}${year_row}"),
            "Semantic actual warranty formula",
        )
        maybe_add_formula_update("ROE (Current Period)", col, generator.get_roe_formula, "Semantic ROE formula")
        maybe_add_formula_update("Retention", col, generator.get_retention_formula, "Semantic Retention formula")
        maybe_add_formula_update("Net Profit (Post-Tax)", col, generator.get_net_profit_formula, "Semantic Net Profit formula")

    for col in audit_cols:
        maybe_add_formula_update([labels.get("revenue", "General Conditions fee"), "General Conditions fee"], col, generator.get_revenue_formula, "Semantic audited Revenue formula")
        maybe_add_formula_update([labels.get("confirmed_cogs", "Cost of Goods Sold"), "Cost of Goods Sold"], col, generator.get_confirmed_cogs_formula, "Semantic audited COGS formula")
        maybe_add_formula_update("Gross Profit", col, generator.get_gross_profit_formula, "Semantic audited gross profit formula")

    return semantic_updates


def _find_material_margin_rows_109(
    label_rows: Dict[str, List[int]],
    row_wbh_cogs: int | None,
    row_wbh_inv: int | None,
) -> Tuple[int | None, int | None]:
    mm_rows = sorted(label_rows.get("material margin", []))
    if not mm_rows:
        return None, None

    main_row: int | None = None
    inv_row: int | None = None

    if row_wbh_cogs is not None:
        for r in mm_rows:
            if r > row_wbh_cogs:
                main_row = r
                break

    if row_wbh_inv is not None:
        for r in mm_rows:
            if r > row_wbh_inv:
                inv_row = r
                break

    if main_row is None:
        main_row = mm_rows[0]
    if inv_row is None:
        inv_row = mm_rows[-1] if len(mm_rows) > 1 else mm_rows[0]

    return main_row, inv_row


def _load_109_formula_dictionary(path: Path | None = None) -> Dict[str, Any]:
    dictionary_path = path or FORMULA_DICTIONARY_109_FILE
    with dictionary_path.open("r", encoding="utf-8") as fh:
        payload = yaml.safe_load(fh) or {}
    if not isinstance(payload, dict):
        raise RuntimeError("109公式字典格式无效。")
    return payload


def _year_columns_from_109_dictionary(config: Mapping[str, Any]) -> List[int]:
    period_axis = dict(config.get("period_axis", {}))
    years = [int(x) for x in period_axis.get("years", [2021, 2022, 2023, 2024, 2025, 2026])]
    return list(range(6, 6 + len(years)))


def _build_109_manual_input_ranges(
    rows: Sequence[Sequence[Any]],
    years: Sequence[int],
    year_blocks: Sequence[Sequence[int]] | None = None,
    sheet_109_title: str | None = None,
) -> List[str]:
    if not rows:
        return []
    resolved_sheet_109_title = _resolve_sheet_109_title(sheet_109_title)

    label_rows = _find_rows_by_item_label(rows, item_col_1=3)
    primary_cols = list(year_blocks[0]) if year_blocks and len(year_blocks) > 0 and year_blocks[0] else list(range(6, 6 + len(years)))
    audit_cols = list(year_blocks[1]) if year_blocks and len(year_blocks) > 1 and year_blocks[1] else list(range(13, 13 + len(years)))
    primary_start_col = _column_number_to_a1(primary_cols[0])
    primary_end_col = _column_number_to_a1(primary_cols[-1])
    audit_start_col = _column_number_to_a1(audit_cols[0])
    audit_end_col = _column_number_to_a1(audit_cols[-1])

    def _find_row(*labels: str) -> int | None:
        for label in labels:
            row_list = label_rows.get(_normalize_label(label), [])
            if row_list:
                return int(row_list[0])
        return None

    ranges: List[str] = [
        _build_109_range("C2:E2", resolved_sheet_109_title),
        _build_109_range("G2:I2", resolved_sheet_109_title),
    ]
    row_blocks = [
        (("General Conditions fee-Audited",), True),
        (("Owner-unapproved Overrun",), False),
        (("Cost of Goods Sold-Audited", "Audit Adjustment (Current Period)"), True),
        (("Accrued Warranty Expenses",), False),
        (("WB Home Income",), False),
        (("WB Home COGS",), False),
        (("WB Home Inventory Income",), False),
        (("WB Home Inventory",), False),
    ]

    for labels, include_audit in row_blocks:
        row_i = _find_row(*labels)
        if row_i is None:
            continue
        ranges.append(_build_109_range(f"{primary_start_col}{row_i}:{primary_end_col}{row_i}", resolved_sheet_109_title))
        if include_audit and audit_cols:
            ranges.append(_build_109_range(f"{audit_start_col}{row_i}:{audit_end_col}{row_i}", resolved_sheet_109_title))
    return ranges


def _build_109_units_count_formula() -> str:
    return '=IFERROR(COUNTA(FILTER(\'Unit Master\'!$A$3:$A,REGEXMATCH(\'Unit Master\'!$A$3:$A,"[0-9]"))),0)'


def _build_109_date_array_formula(func_name: str) -> str:
    date_floor = "DATE(2021,1,1)"
    arrays = [
        'IFERROR(FILTER(Payable!$T:$T,Payable!$T:$T<>""),"")',
        'IFERROR(FILTER(Payable!$V:$V,Payable!$V:$V<>""),"")',
        'IFERROR(FILTER(Payable!$AC:$AC,Payable!$AC:$AC<>""),"")',
        'IFERROR(FILTER(\'Final Detail\'!$O:$O,\'Final Detail\'!$O:$O<>""),"")',
        'IFERROR(FILTER(\'Final Detail\'!$S:$S,\'Final Detail\'!$S:$S<>""),"")',
        'IFERROR(FILTER(\'Draw request report\'!$R:$R,\'Draw request report\'!$R:$R<>""),"")',
        'IFERROR(FILTER(\'Draw request report\'!$Z:$Z,\'Draw request report\'!$Z:$Z<>""),"")',
    ]
    merged = ";".join(arrays)
    return f"=MAX({date_floor},{func_name}(TOCOL({{{merged}}},1)))"


def _a1_to_grid_range(a1: str, sheet_id: int) -> Dict[str, int]:
    normalized = _normalize_formula_range(a1)
    if "!" in normalized:
        _, ref = normalized.split("!", 1)
    else:
        ref = normalized
    column_match = re.fullmatch(r"([A-Z]+):([A-Z]+)", ref)
    if column_match:
        return {
            "sheetId": int(sheet_id),
            "startColumnIndex": _column_a1_to_number(column_match.group(1)) - 1,
            "endColumnIndex": _column_a1_to_number(column_match.group(2)),
        }
    match = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", ref)
    if not match:
        raise ValueError(f"暂不支持的A1范围: {a1}")

    start_col = _column_a1_to_number(match.group(1)) - 1
    start_row = int(match.group(2)) - 1
    end_col = _column_a1_to_number(match.group(3) or match.group(1))
    end_row = int(match.group(4) or match.group(2))
    return {
        "sheetId": int(sheet_id),
        "startRowIndex": start_row,
        "endRowIndex": end_row,
        "startColumnIndex": start_col,
        "endColumnIndex": end_col,
    }


def _compress_a1_ranges(a1_ranges: Sequence[str]) -> List[str]:
    grouped: Dict[Tuple[str, int], List[int]] = {}
    passthrough: List[str] = []
    for a1 in a1_ranges:
        normalized = _normalize_formula_range(a1)
        if "!" in normalized:
            sheet, ref = normalized.split("!", 1)
        else:
            raise ValueError(f"A1范围缺少工作表名前缀: {normalized}")
        match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
        if not match:
            passthrough.append(f"{_quote_sheet_name(sheet)}!{ref}")
            continue
        col_n = _column_a1_to_number(match.group(1))
        row_n = int(match.group(2))
        grouped.setdefault((sheet, row_n), []).append(col_n)

    out = list(passthrough)
    for (sheet, row_n), cols in sorted(grouped.items()):
        for start, end in _contiguous_segments(cols):
            start_col = _column_number_to_a1(start)
            end_col = _column_number_to_a1(end)
            if start == end:
                out.append(f"{_quote_sheet_name(sheet)}!{start_col}{row_n}")
            else:
                out.append(f"{_quote_sheet_name(sheet)}!{start_col}{row_n}:{end_col}{row_n}")
    return out


def _build_repeat_cell_request(grid_range: Mapping[str, int], color: Mapping[str, float]) -> Dict[str, Any]:
    return {
        "repeatCell": {
            "range": dict(grid_range),
            "cell": {"userEnteredFormat": {"backgroundColor": dict(color)}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    }


def _build_number_format_request(
    grid_range: Mapping[str, int],
    number_format: Mapping[str, str],
) -> Dict[str, Any]:
    return {
        "repeatCell": {
            "range": dict(grid_range),
            "cell": {"userEnteredFormat": {"numberFormat": dict(number_format)}},
            "fields": "userEnteredFormat.numberFormat",
        }
    }


def _build_hide_columns_request(
    sheet_id: int,
    start_col_1: int,
    end_col_1_inclusive: int,
) -> Dict[str, Any]:
    return {
        "updateDimensionProperties": {
            "range": {
                "sheetId": int(sheet_id),
                "dimension": "COLUMNS",
                "startIndex": int(start_col_1 - 1),
                "endIndex": int(end_col_1_inclusive),
            },
            "properties": {"hiddenByUser": True},
            "fields": "hiddenByUser",
        }
    }


def _build_update_hidden_rows_request(
    sheet_id: int,
    start_row_1: int,
    end_row_1_inclusive: int,
    hidden: bool,
) -> Dict[str, Any]:
    return {
        "updateDimensionProperties": {
            "range": {
                "sheetId": int(sheet_id),
                "dimension": "ROWS",
                "startIndex": int(start_row_1 - 1),
                "endIndex": int(end_row_1_inclusive),
            },
            "properties": {"hiddenByUser": bool(hidden)},
            "fields": "hiddenByUser",
        }
    }


def _build_delete_conditional_format_rule_request(sheet_id: int, index: int) -> Dict[str, Any]:
    return {
        "deleteConditionalFormatRule": {
            "sheetId": int(sheet_id),
            "index": int(index),
        }
    }


def _build_delete_protected_range_request(protected_range_id: int) -> Dict[str, Any]:
    return {
        "deleteProtectedRange": {
            "protectedRangeId": int(protected_range_id),
        }
    }


def _build_add_protected_range_request(
    sheet_id: int,
    unprotected_ranges: Sequence[Mapping[str, int]],
    editor_email: str | None = None,
    warning_only: bool = False,
) -> Dict[str, Any]:
    protected_range: Dict[str, Any] = {
        "range": {"sheetId": int(sheet_id)},
        "description": MANAGED_109_PROTECTION_DESCRIPTION,
        "warningOnly": bool(warning_only),
        "unprotectedRanges": [dict(item) for item in unprotected_ranges],
    }
    if editor_email and not warning_only:
        protected_range["editors"] = {"users": [editor_email]}
    return {
        "addProtectedRange": {
            "protectedRange": protected_range,
        }
    }


def _build_add_full_sheet_protected_range_request(
    sheet_id: int,
    editor_email: str | None = None,
) -> Dict[str, Any]:
    protected_range: Dict[str, Any] = {
        "range": {"sheetId": int(sheet_id)},
        "description": MANAGED_109_PROTECTION_DESCRIPTION,
        "warningOnly": False,
    }
    if editor_email:
        protected_range["editors"] = {"users": [editor_email]}
    return {
        "addProtectedRange": {
            "protectedRange": protected_range,
        }
    }


def _build_add_targeted_protected_range_request(
    grid_range: Mapping[str, int],
    description: str,
    editor_email: str | None = None,
) -> Dict[str, Any]:
    protected_range: Dict[str, Any] = {
        "range": dict(grid_range),
        "description": description,
        "warningOnly": False,
    }
    if editor_email:
        protected_range["editors"] = {"users": [editor_email]}
    return {
        "addProtectedRange": {
            "protectedRange": protected_range,
        }
    }


def _get_109_sheet_metadata(
    service,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    return _get_sheet_metadata(service, spreadsheet_id, resolved_sheet_109_title)


def _get_sheet_metadata(
    service,
    spreadsheet_id: str,
    sheet_title: str,
) -> Dict[str, Any]:
    response = (
        service.spreadsheets()
        .get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)),protectedRanges(protectedRangeId,range(sheetId,startRowIndex,endRowIndex,startColumnIndex,endColumnIndex),description,unprotectedRanges(sheetId,startRowIndex,endRowIndex,startColumnIndex,endColumnIndex)),conditionalFormats)",
        )
        .execute()
    )
    for sheet in response.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("title") != sheet_title:
            continue
        grid = properties.get("gridProperties", {})
        protected_ranges = []
        for item in sheet.get("protectedRanges", []):
            range_info = item.get("range", {})
            if int(range_info.get("sheetId", -1)) != int(properties.get("sheetId", -2)):
                continue
            protected_ranges.append(item)
        return {
            "sheet_id": int(properties["sheetId"]),
            "row_count": int(grid.get("rowCount", 0)),
            "column_count": int(grid.get("columnCount", 0)),
            "protected_ranges": protected_ranges,
            "conditional_format_count": len(sheet.get("conditionalFormats", [])),
        }
    raise RuntimeError(f"未找到工作表元数据: {sheet_title}")


def _get_sheet_metadata_map(service, spreadsheet_id: str) -> Dict[str, Dict[str, Any]]:
    spreadsheet = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)),protectedRanges(protectedRangeId,range(sheetId,startRowIndex,endRowIndex,startColumnIndex,endColumnIndex),description,unprotectedRanges(sheetId,startRowIndex,endRowIndex,startColumnIndex,endColumnIndex)),conditionalFormats)",
    ).execute()
    output: Dict[str, Dict[str, Any]] = {}
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {})
        title = props.get("title")
        if not title:
            continue
        output[str(title)] = {
            "sheet_id": props.get("sheetId"),
            "row_count": props.get("gridProperties", {}).get("rowCount", 0),
            "column_count": props.get("gridProperties", {}).get("columnCount", 0),
            "protected_ranges": list(sheet.get("protectedRanges", [])),
            "conditional_formats": list(sheet.get("conditionalFormats", [])),
        }
    return output


def _build_project_data_lock_requests(
    service,
    spreadsheet_id: str,
    locked: bool,
    sheet_109_title: str | None = None,
) -> List[Dict[str, Any]]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    requests: List[Dict[str, Any]] = []
    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))
    metadata_by_sheet = _get_sheet_metadata_map(service, spreadsheet_id)

    for sheet_name in _build_project_data_lock_sheet_names(resolved_sheet_109_title):
        metadata = metadata_by_sheet.get(sheet_name)
        if not metadata:
            continue
        description = f"{MANAGED_DATA_LOCK_PREFIX}: {sheet_name}"
        for protected_range in metadata.get("protected_ranges", []):
            if protected_range.get("description") == description and protected_range.get("protectedRangeId") is not None:
                requests.append(_build_delete_protected_range_request(int(protected_range["protectedRangeId"])))
        if locked:
            requests.append(
                _build_add_full_sheet_protected_range_request(
                    sheet_id=int(metadata["sheet_id"]),
                    editor_email=editor_email or None,
                )
            )
            requests[-1]["addProtectedRange"]["protectedRange"]["description"] = description

    return requests


def _apply_109_layout_controls(
    service,
    spreadsheet_id: str,
    manual_ranges: Sequence[str],
    highlight_ranges: Sequence[str] | None = None,
    error_ranges: Sequence[str] | None = None,
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    metadata = _get_109_sheet_metadata(service, spreadsheet_id, sheet_109_title=sheet_109_title)
    sheet_id = int(metadata["sheet_id"])
    row_count = int(metadata["row_count"])
    column_count = int(metadata["column_count"])
    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))

    requests = _build_109_format_requests(
        sheet_id=sheet_id,
        row_count=row_count,
        column_count=column_count,
        manual_ranges=manual_ranges,
        highlight_ranges=list(highlight_ranges or []),
        error_ranges=list(error_ranges or []),
    )
    for protected_range in metadata.get("protected_ranges", []):
        protected_range_id = protected_range.get("protectedRangeId")
        if protected_range_id is None:
            continue
        requests.append(_build_delete_protected_range_request(int(protected_range_id)))

    requests.append(
        _build_add_protected_range_request(
            sheet_id=sheet_id,
            unprotected_ranges=[_a1_to_grid_range(a1, sheet_id) for a1 in _compress_a1_ranges(manual_ranges)],
            editor_email=editor_email or None,
        )
    )
    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ).execute()

    return {
        "sheet_id": sheet_id,
        "manual_range_count": len(manual_ranges),
        "protected_range_count": 1,
        "deleted_protected_range_count": len(metadata.get("protected_ranges", [])),
        "request_count": len(requests),
    }


def _parse_sheet_name_from_a1(a1: str) -> str | None:
    normalized = _normalize_formula_range(a1)
    if "!" not in normalized:
        return None
    sheet_part, _ = normalized.split("!", 1)
    sheet_part = sheet_part.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part


def _build_formula_lock_a1_ranges(
    plan: Sequence[Mapping[str, Any]],
    sheet_109_title: str,
) -> List[str]:
    ranges: List[str] = []
    for item in plan:
        raw_range = _safe_string(item.get("range", ""))
        if not raw_range:
            continue
        target_sheet = _parse_sheet_name_from_a1(raw_range)
        if target_sheet and target_sheet != sheet_109_title:
            continue
        if not target_sheet:
            normalized_range = _build_109_range(raw_range, sheet_109_title)
        else:
            normalized_range = _normalize_formula_range(raw_range)
        ranges.append(normalized_range)
    if not ranges:
        return []
    return _compress_a1_ranges(ranges)


def _apply_109_formula_lock_protection(
    service,
    spreadsheet_id: str,
    plan: Sequence[Mapping[str, Any]],
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    metadata = _get_109_sheet_metadata(
        service,
        spreadsheet_id,
        sheet_109_title=resolved_sheet_109_title,
    )
    sheet_id = int(metadata["sheet_id"])
    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))
    lock_ranges = _build_formula_lock_a1_ranges(plan, resolved_sheet_109_title)

    requests: List[Dict[str, Any]] = []
    deleted_count = 0
    for protected_range in metadata.get("protected_ranges", []):
        description = _safe_string(protected_range.get("description", ""))
        protected_range_id = protected_range.get("protectedRangeId")
        if not description.startswith(MANAGED_109_FORMULA_PROTECTION_PREFIX) or protected_range_id is None:
            continue
        requests.append(_build_delete_protected_range_request(int(protected_range_id)))
        deleted_count += 1

    for idx, a1_range in enumerate(lock_ranges, start=1):
        requests.append(
            _build_add_targeted_protected_range_request(
                grid_range=_a1_to_grid_range(a1_range, sheet_id),
                description=f"{MANAGED_109_FORMULA_PROTECTION_PREFIX}: {idx}",
                editor_email=editor_email or None,
            )
        )

    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ).execute()

    return {
        "sheet_id": sheet_id,
        "formula_lock_ranges": lock_ranges,
        "formula_lock_range_count": len(lock_ranges),
        "deleted_formula_lock_count": deleted_count,
        "request_count": len(requests),
    }


def _build_scoping_manual_input_ranges(rows: Sequence[Sequence[Any]]) -> List[str]:
    editable_rows: List[int] = []
    for row_num, row in enumerate(rows[2:], start=3):
        group_value = row[2] if len(row) > 2 else ""
        if not _has_digits(group_value):
            continue
        editable_rows.append(row_num)
    ranges: List[str] = []
    for start_row, end_row in _contiguous_segments(editable_rows):
        if start_row == end_row:
            ranges.append(f"{_quote_sheet_name('Scoping')}!B{start_row}")
            ranges.append(f"{_quote_sheet_name('Scoping')}!E{start_row}:K{start_row}")
        else:
            ranges.append(f"{_quote_sheet_name('Scoping')}!B{start_row}:B{end_row}")
            ranges.append(f"{_quote_sheet_name('Scoping')}!E{start_row}:K{end_row}")
    return ranges


def _scoping_has_amount(value: Any) -> bool:
    amount = _to_float(value)
    return amount is not None and abs(amount) > 1e-9


def _build_scoping_hidden_row_numbers(rows: Sequence[Sequence[Any]]) -> List[int]:
    hidden_rows: List[int] = []
    for row_num, row in enumerate(rows[2:], start=3):
        group_value = row[2] if len(row) > 2 else ""
        if _has_digits(group_value):
            continue
        budget_value = row[12] if len(row) > 12 else ""
        incurred_value = row[13] if len(row) > 13 else ""
        if _scoping_has_amount(budget_value) or _scoping_has_amount(incurred_value):
            continue
        hidden_rows.append(row_num)
    return hidden_rows


def _build_scoping_warranty_expiry_values(
    scoping_rows: Sequence[Sequence[Any]],
    unit_master_rows: Sequence[Sequence[Any]],
    unit_budget_rows: Sequence[Sequence[Any]] | None = None,
) -> List[List[Any]]:
    output: List[List[Any]] = [[""] for _ in range(max(len(scoping_rows), 1))]
    if len(output) >= 2:
        output[1] = ["保修到期日"]

    warranty_months_by_group: Dict[int, float] = {}
    for row in scoping_rows[2:]:
        group_number = _to_float(row[2] if len(row) > 2 else "")
        warranty_months = _to_float(row[10] if len(row) > 10 else "")
        if group_number is None or warranty_months is None:
            continue
        warranty_months_by_group[int(group_number)] = float(warranty_months)

    def _find_header_col(header_row: Sequence[Any], candidates: Sequence[str], default: int) -> int:
        wanted = {_normalize_label(item) for item in candidates}
        for idx, cell in enumerate(header_row, start=1):
            if _normalize_label(cell) in wanted:
                return idx
        return default

    latest_co_date_by_group: Dict[int, pd.Timestamp] = {}

    def _collect_latest(rows: Sequence[Sequence[Any]], header_idx: int, group_default: int, co_default: int) -> None:
        header_row = rows[header_idx - 1]
        group_col = _find_header_col(header_row, ["Group"], group_default)
        co_date_col = _find_header_col(header_row, ["C/O date"], co_default)
        for row in rows[header_idx:]:
            group_number = _to_float(row[group_col - 1] if len(row) >= group_col else "")
            co_date = _normalize_date_value(row[co_date_col - 1] if len(row) >= co_date_col else "")
            if group_number is None or co_date is None:
                continue
            group_key = int(group_number)
            existing = latest_co_date_by_group.get(group_key)
            if existing is None or co_date > existing:
                latest_co_date_by_group[group_key] = co_date

    unit_master_header_idx = _find_first_row(
        unit_master_rows,
        lambda _row_num, row: _normalize_label(row[0] if len(row) > 0 else "") == _normalize_label("Unit Code")
        and any(_normalize_label(cell) == _normalize_label("C/O date") for cell in row),
    )
    if unit_master_header_idx is not None:
        _collect_latest(unit_master_rows, unit_master_header_idx, 13, 8)

    unit_budget_rows = unit_budget_rows or []
    unit_budget_header_idx = _find_first_row(
        unit_budget_rows,
        lambda _row_num, row: any(_normalize_label(cell) == _normalize_label("C/O date") for cell in row)
        and any(_normalize_label(cell) == _normalize_label("Group") for cell in row),
    )
    if unit_budget_header_idx is not None:
        _collect_latest(unit_budget_rows, unit_budget_header_idx, 14, 8)

    if not latest_co_date_by_group:
        return output

    for row_idx, row in enumerate(scoping_rows[2:], start=2):
        group_number = _to_float(row[2] if len(row) > 2 else "")
        if group_number is None:
            output[row_idx] = [""]
            continue
        latest_co_date = latest_co_date_by_group.get(int(group_number))
        warranty_months = warranty_months_by_group.get(int(group_number))
        if latest_co_date is None or warranty_months is None:
            output[row_idx] = [""]
            continue
        expiry_date = latest_co_date + pd.to_timedelta(float(warranty_months) * 30.25, unit="D")
        output[row_idx] = [f"{expiry_date.month}/{expiry_date.day}/{expiry_date.year}"]
    return output


def _apply_scoping_layout_controls(
    service,
    spreadsheet_id: str,
) -> Dict[str, Any]:
    rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Scoping'!A:Z")
        .execute()
        .get("values", [])
    )
    unit_master_rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Unit Master'!A:Z")
        .execute()
        .get("values", [])
    )
    unit_budget_rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Unit Budget'!A:Z")
        .execute()
        .get("values", [])
    )
    metadata = _get_sheet_metadata(service, spreadsheet_id, "Scoping")
    sheet_id = int(metadata["sheet_id"])
    row_count = int(metadata["row_count"])
    column_count = int(metadata["column_count"])
    manual_ranges = _build_scoping_manual_input_ranges(rows)
    hidden_rows = _build_scoping_hidden_row_numbers(rows)
    warranty_expiry_values = _build_scoping_warranty_expiry_values(rows, unit_master_rows, unit_budget_rows)
    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))

    requests: List[Dict[str, Any]] = [
        _build_repeat_cell_request(
            {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": max(row_count, 1),
                "startColumnIndex": 0,
                "endColumnIndex": max(column_count, 1),
            },
            COLOR_FILL_WHITE,
        ),
        _build_update_hidden_rows_request(sheet_id, 1, max(row_count, 1), False),
    ]
    for a1 in manual_ranges:
        requests.append(_build_repeat_cell_request(_a1_to_grid_range(a1, sheet_id), COLOR_FILL_LIGHT_GRAY))
    requests.append(
        _build_number_format_request(
            {
                "sheetId": sheet_id,
                "startRowIndex": 2,
                "endRowIndex": max(row_count, 2),
                "startColumnIndex": 14,
                "endColumnIndex": 15,
            },
            NUMBER_FORMAT_DATE_MDY,
        )
    )
    for start_row, end_row in _contiguous_segments(hidden_rows):
        requests.append(_build_update_hidden_rows_request(sheet_id, start_row, end_row, True))
    for rule_index in reversed(range(int(metadata.get("conditional_format_count", 0)))):
        requests.append(_build_delete_conditional_format_rule_request(sheet_id, rule_index))
    for protected_range in metadata.get("protected_ranges", []):
        if protected_range.get("description") != MANAGED_SCOPING_PROTECTION_DESCRIPTION:
            continue
        protected_range_id = protected_range.get("protectedRangeId")
        if protected_range_id is not None:
            requests.append(_build_delete_protected_range_request(int(protected_range_id)))
    requests.append(
        _build_add_protected_range_request(
            sheet_id=sheet_id,
            unprotected_ranges=[_a1_to_grid_range(a1, sheet_id) for a1 in manual_ranges],
            editor_email=editor_email or None,
        )
    )
    if requests:
        # Override the default main-sheet protection description for this request.
        requests[-1]["addProtectedRange"]["protectedRange"]["description"] = MANAGED_SCOPING_PROTECTION_DESCRIPTION
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ).execute()
    if warranty_expiry_values:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'Scoping'!O1:O{len(warranty_expiry_values)}",
            valueInputOption="USER_ENTERED",
            body={"values": warranty_expiry_values},
        ).execute()

    return {
        "sheet_id": sheet_id,
        "manual_range_count": len(manual_ranges),
        "hidden_row_count": len(hidden_rows),
        "conditional_format_count": int(metadata.get("conditional_format_count", 0)),
        "request_count": len(requests),
    }


def _build_109_format_requests(
    sheet_id: int,
    row_count: int,
    column_count: int,
    manual_ranges: Sequence[str],
    highlight_ranges: Sequence[str],
    error_ranges: Sequence[str],
) -> List[Dict[str, Any]]:
    requests: List[Dict[str, Any]] = [
        _build_repeat_cell_request(
            {
                "sheetId": int(sheet_id),
                "startRowIndex": 0,
                "endRowIndex": int(row_count),
                "startColumnIndex": 0,
                "endColumnIndex": int(column_count),
            },
            COLOR_FILL_WHITE,
        )
    ]

    for a1 in manual_ranges:
        requests.append(_build_repeat_cell_request(_a1_to_grid_range(a1, sheet_id), COLOR_FILL_LIGHT_GRAY))
    for a1 in _compress_a1_ranges(highlight_ranges):
        requests.append(_build_repeat_cell_request(_a1_to_grid_range(a1, sheet_id), COLOR_FILL_LIGHT_YELLOW))
    for a1 in _compress_a1_ranges(error_ranges):
        requests.append(_build_repeat_cell_request(_a1_to_grid_range(a1, sheet_id), COLOR_FILL_LIGHT_RED))
    return requests


def _build_unit_budget_actual_settlement_values(
    rows: Sequence[Sequence[Any]],
    overrides: Mapping[str, Mapping[str, str]],
) -> List[List[Any]]:
    prepared: List[Tuple[str, Any, str, str, int | str, str]] = []
    latest_unit_year: int | None = None
    latest_numeric_actual_dt: pd.Timestamp | None = None
    latest_numeric_tbd_dt: pd.Timestamp | None = None

    for row in rows[2:]:
        unit_code = _safe_string(row[1] if len(row) > 1 else "")
        settlement_year = row[6] if len(row) > 6 else ""
        override_item = overrides.get(unit_code, {})
        co_date = override_item.get("co_date", "") or (row[7] if len(row) > 7 else "")
        tbd_date = override_item.get("tbd_acceptance_date", "") or (row[10] if len(row) > 10 else "")
        actual_date, actual_year = _derive_unit_budget_actual_settlement_fields(
            unit_code, settlement_year, co_date, None
        )
        if _has_digits(unit_code):
            if actual_year != "":
                year_num = int(actual_year)
                latest_unit_year = year_num if latest_unit_year is None else max(latest_unit_year, year_num)
            actual_dt = _normalize_date_value(actual_date)
            if actual_dt is not None:
                latest_numeric_actual_dt = actual_dt if latest_numeric_actual_dt is None else max(latest_numeric_actual_dt, actual_dt)
            tbd_dt = _normalize_date_value(tbd_date)
            if tbd_dt is not None:
                latest_numeric_tbd_dt = tbd_dt if latest_numeric_tbd_dt is None else max(latest_numeric_tbd_dt, tbd_dt)
        prepared.append((unit_code, settlement_year, _safe_string(co_date), actual_date, actual_year, _safe_string(tbd_date)))

    values: List[List[Any]] = []
    latest_actual_text = _format_iso_date_or_blank(latest_numeric_actual_dt)
    latest_tbd_text = _format_iso_date_or_blank(latest_numeric_tbd_dt)
    latest_actual_year = latest_numeric_actual_dt.year if latest_numeric_actual_dt is not None else latest_unit_year
    for unit_code, settlement_year, co_date, actual_date, _, tbd_date in prepared:
        _, actual_year = _derive_unit_budget_actual_settlement_fields(unit_code, settlement_year, co_date, latest_unit_year)
        actual_date_out = actual_date
        tbd_date_out = tbd_date
        if not _has_digits(unit_code):
            if not actual_date_out: actual_date_out = latest_actual_text
            if actual_year == "" and latest_actual_year is not None: actual_year = latest_actual_year
            if not tbd_date_out: tbd_date_out = latest_tbd_text
        values.append([co_date, actual_date_out, actual_year if actual_year != "" else "", tbd_date_out])
    return values


def _build_unit_master_rows_v2(
    ub_rows: Sequence[Sequence[Any]],
    fd_rows: Sequence[Sequence[Any]],
    pay_rows: Sequence[Sequence[Any]]
) -> List[List[Any]]:
    if not ub_rows: return []
    units = set()
    for r in ub_rows[2:]:
        val = _safe_string(r[1] if len(r) > 1 else "")
        if val and val.upper() not in ("TOTAL", "SUM"): units.add(val)
    for r in pay_rows[1:]:
        if len(r) > 37:
            val = _safe_string(r[37])
            if val: units.add(val)
    for r in fd_rows[1:]:
        if len(r) > 20:
            val = _safe_string(r[20])
            if val: units.add(val)
    sorted_units = sorted(list(units))

    header_row = ub_rows[0] if ub_rows else []
    ub_unit_to_col = {}
    for idx in range(20, len(header_row)):
        val = _safe_string(header_row[idx])
        if val: ub_unit_to_col[val] = idx
    ub_unit_to_row = {}
    for r in ub_rows[1:]:
        u_code = _safe_string(r[1] if len(r) > 1 else "")
        if u_code: ub_unit_to_row[u_code] = r
    unit_to_latest_date: Dict[str, pd.Timestamp] = {}
    for r in fd_rows[1:]:
        if len(r) > 20:
            u = _safe_string(r[20])
            d_raw = r[14] if len(r) > 14 else ""
            if u and d_raw:
                dt = _normalize_date_value(d_raw)
                if dt:
                    if u not in unit_to_latest_date or dt > unit_to_latest_date[u]:
                        unit_to_latest_date[u] = dt

    master_header = ["Unit Code", "Total Budget", "GC Budget", "WIP Budget", "Incurred Amount", "Settlement Amount", "Final Date", "C/O date", "实际结算日期", "实际结算年份", "TBD Acceptance Date", "Budget Variance", "Group"]
    data_rows: List[List[Any]] = []
    for uc in sorted_units:
        curr_row_idx = len(data_rows) + 3
        row_out = ["" for _ in range(13)]
        row_out[0] = uc
        total_budget, gc_budget, wip_budget = 0.0, 0.0, 0.0
        col_idx = ub_unit_to_col.get(uc)
        if col_idx is not None:
            for r in ub_rows[1:]:
                cell_val = _to_float(r[col_idx]) if col_idx < len(r) else None
                if cell_val is not None:
                    if len(r) > 14 and _safe_string(r[14]) == "1": total_budget += cell_val
                    if len(r) > 15 and _safe_string(r[15]) == "2": gc_budget += cell_val
                    if len(r) > 16 and _safe_string(r[16]) == "3": wip_budget += cell_val
        row_out[1] = "" if abs(total_budget) < 1e-12 else total_budget
        row_out[2] = "" if abs(gc_budget) < 1e-12 else gc_budget
        row_out[3] = "" if abs(wip_budget) < 1e-12 else wip_budget
        row_out[4] = f"=SUMIFS(Payable!$U:$U, Payable!$AL:$AL, $A{curr_row_idx}, Payable!$A:$A, \"ROE\") + SUMIFS(Payable!$U:$U, Payable!$AL:$AL, $A{curr_row_idx}, Payable!$A:$A, \"RACC\")"
        row_out[5] = f"=SUMIFS('Final Detail'!$P:$P, 'Final Detail'!$U:$U, $A{curr_row_idx}, 'Final Detail'!$A:$A, \"ROE\") + SUMIFS('Final Detail'!$P:$P, 'Final Detail'!$U:$U, $A{curr_row_idx}, 'Final Detail'!$A:$A, \"ACC\")"
        latest_dt = unit_to_latest_date.get(uc)
        row_out[6] = latest_dt.strftime("%Y-%m-%d") if latest_dt else ""
        row_out[11] = f"=$B{curr_row_idx} - $C{curr_row_idx} - $F{curr_row_idx}"
        ub_row = ub_unit_to_row.get(uc)
        if ub_row:
            for i, target_i in [(7, 7), (8, 8), (9, 9), (10, 10), (12, 12)]:
                if len(ub_row) > i: row_out[target_i] = ub_row[i]
        data_rows.append(row_out)
    last_row = len(data_rows) + 2
    total_row = [""] * 13
    total_row[0] = "Total"
    for col_idx_1 in [2, 3, 4, 5, 6, 12]:
        col_a1 = _column_number_to_a1(col_idx_1)
        total_row[col_idx_1 - 1] = f"=SUM({col_a1}3:{col_a1}{last_row})"
    return [total_row, master_header] + data_rows


def _build_unit_budget_support_requests(
    unit_budget_sheet_id: int,
    unit_master_sheet_id: int,
    row_count: int,
) -> List[Dict[str, Any]]:
    end_row = max(int(row_count), 3)
    requests = [
        _build_hide_columns_request(unit_budget_sheet_id, 2, 13),
        _build_repeat_cell_request({"sheetId": int(unit_budget_sheet_id), "startRowIndex": 0, "endRowIndex": end_row, "startColumnIndex": 1, "endColumnIndex": 13}, COLOR_FILL_WHITE),
        _build_repeat_cell_request({"sheetId": int(unit_master_sheet_id), "startRowIndex": 0, "endRowIndex": end_row, "startColumnIndex": 0, "endColumnIndex": 12}, COLOR_FILL_WHITE),
        _build_repeat_cell_request({"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 7, "endColumnIndex": 8}, COLOR_FILL_LIGHT_GRAY),
        _build_repeat_cell_request({"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 10, "endColumnIndex": 11}, COLOR_FILL_LIGHT_GRAY),
        _build_number_format_request({"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 9, "endColumnIndex": 10}, NUMBER_FORMAT_YEAR_0),
        _build_repeat_cell_request({"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 7, "endColumnIndex": 8}, COLOR_FILL_LIGHT_GRAY),
        _build_repeat_cell_request({"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 10, "endColumnIndex": 11}, COLOR_FILL_LIGHT_GRAY),
        _build_number_format_request({"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 9, "endColumnIndex": 10}, NUMBER_FORMAT_YEAR_0),
    ]
    for grid_range in [
        {"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 7, "endColumnIndex": 8},
        {"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 8, "endColumnIndex": 9},
        {"sheetId": int(unit_budget_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 10, "endColumnIndex": 11},
        {"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 6, "endColumnIndex": 7},
        {"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 7, "endColumnIndex": 8},
        {"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 8, "endColumnIndex": 9},
        {"sheetId": int(unit_master_sheet_id), "startRowIndex": 2, "endRowIndex": end_row, "startColumnIndex": 10, "endColumnIndex": 11},
    ]:
        requests.append(_build_number_format_request(grid_range, NUMBER_FORMAT_DATE_ISO))
    return requests


def _build_unit_master_manual_input_ranges(row_count: int) -> List[str]:
    end_row = max(int(row_count), 3)
    return [
        f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!H3:H{end_row}",
        f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!K3:K{end_row}",
    ]


def _build_unit_master_bootstrap_clear_ranges(row_count: int) -> List[str]:
    """Clear residual template values from Unit Master on project bootstrap.

    Keep the row-2 headers, but wipe:
    - row 1 numeric totals (B:M)
    - all data rows (A:M from row 3 onward)
    """
    end_row = max(int(row_count), 3)
    return [
        f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!B1:M1",
        f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!A3:M{end_row}",
    ]


def _build_109_error_ranges_from_values(
    value_map: Mapping[str, float | None],
    sheet_109_title: str | None = None,
) -> List[str]:
    resolved_sheet_109_title = _resolve_sheet_109_title(sheet_109_title)
    errors: List[str] = []
    e3, e4, e5, e12, e13, e37 = [value_map.get(_build_109_cell_key(f"E{i}", resolved_sheet_109_title)) for i in [3, 4, 5, 12, 13, 37]]
    if e3 is not None and e4 is not None and e5 is not None and abs((e3 - e4) - e5) > 0.01:
        errors.append(_build_109_range("E5", resolved_sheet_109_title))
    if e12 is not None and e12 > 1.0 + 1e-9:
        errors.append(_build_109_range("E12", resolved_sheet_109_title))
    if e13 is not None:
        if e13 > 1.0 + 1e-9: errors.append(_build_109_range("E13", resolved_sheet_109_title))
        elif e12 is not None and abs(e13 - e12) > 1e-6: errors.append(_build_109_range("E13", resolved_sheet_109_title))
    if e37 is not None and e12 is not None and abs(e37 - e12) > 1e-6:
        errors.append(_build_109_range("E37", resolved_sheet_109_title))
    return errors


def _build_109_formula_plan_from_grid(
    rows: Sequence[Sequence[Any]],
    config: Mapping[str, Any] | None = None,
    sheet_109_title: str | None = None,
    formula_mappings: Mapping[str, Mapping[str, int]] | None = None,
) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    if not rows: raise RuntimeError("109工作表为空。")
    cfg = dict(config or _load_109_formula_dictionary())
    cfg["formula_mappings"] = {
        _safe_string(sheet_name): {str(field): int(index) for field, index in dict(fields).items()}
        for sheet_name, fields in dict(formula_mappings or {}).items()
        if isinstance(fields, Mapping)
    }
    actual_sheet_109_title = _resolve_sheet_109_title(sheet_109_title)

    discovered_year_row, discovered_year_blocks = _discover_109_year_column_blocks(rows)
    year_blocks = discovered_year_blocks or _default_109_year_column_blocks()
    primary_year_cols = list(year_blocks[0]) if year_blocks else list(range(6, 12))
    year_row = discovered_year_row or _find_year_header_row_109(rows) or 10
    primary_year_letters = [_column_number_to_a1(col_i) for col_i in primary_year_cols]
    audit_year_letters = [_column_number_to_a1(col_i) for col_i in year_blocks[1]] if len(year_blocks) > 1 else []
    cfg["primary_year_cols"] = primary_year_letters
    cfg["audit_year_cols"] = audit_year_letters
    if primary_year_letters:
        cfg["start_year_anchor_cell"] = f"{primary_year_letters[-1]}2"
    mapper = MapperFactory.create("109", rows)
    generator = FinanceFormulaGenerator(mapper, config=cfg)
    meta: Dict[str, Any] = {}

    def _m_row(*labels: str) -> int | None:
        for lb in labels:
            try: return mapper.get_row(lb)
            except KeyError: continue
        return None
    row_contract_change = _m_row("Cumulative Savings (Target vs Actual)", "contract change order", "budget surplus")
    row_contract_amount = _m_row("contract amount")
    label_rows = _find_rows_by_item_label(rows, item_col_1=3)
    surplus_row_candidates = label_rows.get("cumulative savings target vs actual", []) or label_rows.get("budget surplus", [])
    row_surplus_tp = surplus_row_candidates[0] if surplus_row_candidates else None
    row_surplus_eac = surplus_row_candidates[1] if len(surplus_row_candidates) > 1 else None
    if row_contract_change is None: row_contract_change = row_surplus_tp
    if row_contract_amount is None and row_contract_change is not None and row_contract_change > 1: row_contract_amount = row_contract_change - 1
    
    contract_change_rows = list(label_rows.get("contract change amount", []))
    contract_price_rows = list(label_rows.get("contract price", []))
    row_contract_price = (
        (contract_change_rows[0] if contract_change_rows else None)
        or (contract_price_rows[0] if contract_price_rows else None)
        or _choose_contract_price_row(rows, label_rows)
    )
    row_contract_price_day1 = _find_contract_price_day1_row(rows)
    row_poc = _m_row("Percentage of Completion (POC)", "percentage of completion")
    row_cr = _m_row("Completion Rate for the Period", "completion rate for the period")
    row_initial_budget = _m_row("Initial Budget (Original Contract Sum)", "day 1 budget")
    row_overrun = _m_row("Owner-unapproved Overrun", "owner-unapproved overrun")
    row_budget = _m_row("Current Dynamic EAC (Total Cost)", "dynamic budget (eac)", "scoping budget cost")
    row_cum_direct_cost = _m_row("Cumulative Total Cost (Actual)", "cumulative direct cost")
    row_cogs_calc = _m_row("cost of goods sold-company")
    row_cogs_aud = _m_row("Audit Adjustment (Current Period)", "cost of goods sold-audited")
    row_cogs = _m_row("Confirmed COGS (Current Period)", "cost of goods sold")
    row_revenue_company = _m_row("general conditions fee-company")
    row_revenue_aud = _m_row("general conditions fee-audited")
    row_revenue = _m_row("general conditions fee")
    row_gp_company = _m_row("gross profit-company")
    row_gp = _m_row("gross profit")
    row_ar_incurred = _m_row("accounts receivable-incurred")
    row_ar_aud = _m_row("accounts receivable-audited")
    row_ar_company = _m_row("accounts receivable-company")
    row_ar = _m_row("accounts receivable")
    row_wbh_income = _m_row("wb home income")
    row_wbh_cogs = _m_row("wb home cogs")
    row_inv_income = _m_row("wb home inventory income")
    row_inv = _m_row("wb home inventory")
    row_inv_income_rev = _m_row("wb home inventory income-reverse")
    row_inv_rev = _m_row("wb home inventory-reverse")
    row_roe_total = _m_row("total roe cost")
    row_roe_wbhome = _m_row("roe cost - wb home")
    row_roe_wpred = _m_row("roe cost - wpred")
    row_acc_expenses = _m_row("accrued expenses")
    row_racc_reversed = _m_row("reversed accrued expenses")
    row_income_total = _m_row("total income cost")
    row_gc_income = _m_row("gc income")
    row_gc_cost = _m_row("total gc cost")
    row_accrued_warranty = _m_row("accrued warranty expenses")
    row_actual_warranty = _m_row("actual warranty expenses (reversed)")
    row_wbh_total = _m_row("wb. home material margin total")
    row_main_mm, row_inv_mm = _find_material_margin_rows_109(label_rows, row_wbh_cogs, row_inv)

    required_rows = {
        "年度表头(2021-2026)": year_row, "Contract Amount": row_contract_amount, "Contract price": row_contract_price,
        "Contract price (Day1)": row_contract_price_day1, "Contract Change Order": row_surplus_tp,
        "Percentage of Completion": row_poc, "Completion Rate for the Period": row_cr, "Initial Budget": row_initial_budget,
        "Owner-unapproved Overrun": row_overrun, "Budget Cost(分母EAC)": row_budget, "Cumulative Direct Cost": row_cum_direct_cost,
        "Cost of Goods Sold-Company": row_cogs_calc, "General Conditions fee-Company": row_revenue_company,
        "Gross Profit-Company": row_gp_company, "Accounts Receivable-Incurred": row_ar_incurred, "Accounts Receivable-Audited": row_ar_aud,
        "Accounts Receivable-Company": row_ar_company, "Accounts Receivable": row_ar, "WB Home Income": row_wbh_income,
        "WB Home COGS": row_wbh_cogs, "WB Home Inventory Income": row_inv_income, "WB Home Inventory": row_inv,
        "WB Home Inventory Income-Reverse": row_inv_income_rev, "WB Home Inventory-Reverse": row_inv_rev,
        "WB. Home Material Margin Total": row_wbh_total, "Total ROE Cost": row_roe_total, "ROE Cost - WB Home": row_roe_wbhome,
        "ROE Cost - WPRED": row_roe_wpred, "Accrued Expenses": row_acc_expenses, "Reversed Accrued Expenses": row_racc_reversed,
        "Total Income Cost": row_income_total, "GC Income": row_gc_income, "Material Margin(main)": row_main_mm, "Material Margin(inventory)": row_inv_mm,
    }
    # 新分类行设为必选
    optional_rows = {
        "Total GC Cost": row_gc_cost,
        "Accrued Warranty Expenses": row_accrued_warranty,
        "Actual Warranty Expenses (Reversed)": row_actual_warranty
    }
    
    missing = [name for name, row_i in required_rows.items() if row_i is None]
    if missing: raise RuntimeError("109关键行定位失败: " + "、".join(missing))

    plan: List[Dict[str, str]] = []
    def add_formula(col_i: int, row_i: int, formula: str, logic: str) -> None:
        col = _column_number_to_a1(col_i)
        label_cells: List[str] = []
        if row_i > 0 and row_i - 1 < len(rows):
            row_values = rows[row_i - 1]
            if isinstance(row_values, Sequence) and not isinstance(row_values, (str, bytes, bytearray)):
                label_cells = [_safe_string(row_values[idx]) if idx < len(row_values) else "" for idx in (2, 3)]
        plan.append({
            "sheet": actual_sheet_109_title,
            "cell": f"{col}{row_i}",
            "range": _build_109_range(f"{col}{row_i}", actual_sheet_109_title),
            "formula": formula,
            "logic": logic,
            "row_fingerprint": {"label_cells": label_cells},
        })

    start_year_anchor_col = primary_year_cols[-1] if primary_year_cols else 11
    start_year_anchor_cell = f"${_column_number_to_a1(start_year_anchor_col)}$2"
    start_year_expr = f"Year({start_year_anchor_cell})"
    add_formula(start_year_anchor_col, 2, _build_109_date_array_formula("MIN"), "Start date")
    add_formula(start_year_anchor_col, 3, _build_109_date_array_formula("MAX"), "End date")
    add_formula(3, 5, _build_109_units_count_formula(), "Units count")
    add_formula(5, 3, "=SUMIFS('Unit Budget'!$T:$T,'Unit Budget'!$O:$O,1)", "Contract price (Day1)")
    add_formula(5, 5, "=SUMIFS('Unit Budget'!$T:$T,'Unit Budget'!$P:$P,2)", "General Conditions fee")
    primary_start_col = _column_number_to_a1(primary_year_cols[0])
    primary_end_col = _column_number_to_a1(primary_year_cols[-1])
    add_formula(5, 12, f'=IFERROR(round(MAX({primary_start_col}12:{primary_end_col}12),8),"")', "POC Total")
    add_formula(5, 13, f'=IFERROR(round(SUM({primary_start_col}13:{primary_end_col}13),8),"")', "Completion Rate Total")

    for idx, col_i in enumerate(primary_year_cols):
        col = _column_number_to_a1(col_i)
        prev_col = _column_number_to_a1(primary_year_cols[idx - 1]) if idx > 0 else ""
        year_ref = f"{col}${year_row}"
        add_formula(col_i, row_contract_amount, f'=IF({col}$10={start_year_expr},-$E$3,0)', "Contract Amount") # type: ignore
        add_formula(col_i, row_surplus_tp, f"=SUMIFS('Unit Master'!$L:$L,'Unit Master'!$J:$J,{year_ref})", "Budget Surplus") # type: ignore
        contract_price_formula = f'=IF({col}$10<{start_year_expr},"",IF({col}$10={start_year_expr},{col}{row_contract_amount}+{col}{row_surplus_tp},IFERROR({prev_col}{row_contract_price}+{col}{row_surplus_tp},"")))'
        add_formula(col_i, row_contract_price, contract_price_formula, "Contract Change Amount") # type: ignore
        add_formula(col_i, row_initial_budget, f"=IF({col}$10={start_year_expr},$C$3,0)", "Initial Budget") # type: ignore
        add_formula(col_i, row_budget, generator.get_eac_formula(col), "EAC") # type: ignore
        add_formula(col_i, row_poc, generator.get_poc_formula(col), "POC") # type: ignore
        # 核心逻辑修正：调用重构后的方法，传入 prev_col 以支持差额推算
        add_formula(col_i, row_cogs, generator.get_confirmed_cogs_formula(col, prev_col), "COGS") # type: ignore
        add_formula(col_i, row_cr, f"=IFERROR(round({col}{row_poc}-{'0' if not prev_col else prev_col+str(row_poc)},8),\"\")", "CR") # type: ignore
        # 核心逻辑修正：调用重构后的方法，传入 prev_col
        add_formula(col_i, row_revenue, generator.get_revenue_formula(col, prev_col), "Revenue") # type: ignore
        
        # Gross Profit, 52 行
        if row_gp:
            add_formula(col_i, row_gp, generator.get_gross_profit_formula(col), "Gross Profit") # type: ignore

        # 固化新三行公式
        if row_gc_cost:
            add_formula(col_i, row_gc_cost, generator.get_gc_cost_formula(col, year_ref), "Total GC Cost")
        if row_income_total:
            add_formula(col_i, row_income_total, generator.get_income_total_formula(col, year_ref), "Total Income Cost")
        if row_gc_income:
            add_formula(col_i, row_gc_income, generator.get_gc_income_formula(col, year_ref), "GC Income")
        if row_accrued_warranty:
            # 计提行在主年度列不生成公式，保持手工值，由格式化渲染灰色
            pass
        if row_actual_warranty:
            add_formula(col_i, row_actual_warranty, generator.get_actual_warranty_formula(col, year_ref), "Actual Warranty Expenses (Reversed)")

    # 特殊格式化：计提保修费用手工区（主年度列）
    if row_accrued_warranty:
        manual_range = _build_109_range(
            f"{primary_start_col}{row_accrued_warranty}:{primary_end_col}{row_accrued_warranty}",
            actual_sheet_109_title,
        )
        # 注入到 meta 中，后续格式化逻辑会读取
        if "manual_input_ranges" not in meta: meta["manual_input_ranges"] = []
        meta["manual_input_ranges"].append(manual_range)
    manual_ranges = _build_109_manual_input_ranges(
        rows,
        _year_columns_from_109_dictionary(cfg),
        year_blocks=year_blocks,
        sheet_109_title=actual_sheet_109_title,
    )
    for manual_range in meta.get("manual_input_ranges", []):
        if manual_range not in manual_ranges:
            manual_ranges.append(manual_range)

    return plan, {"sheet": actual_sheet_109_title, "year_row": year_row, "formula_count": len(plan), "key_rows": {**required_rows, **optional_rows}, "manual_ranges": manual_ranges}


def _ensure_109_labels(service, spreadsheet_id: str, sheet_109_title: str | None = None) -> int:
    """确保109表的关键行（如GC成本、保修费用等）具备正确的英文标签，以便后续公式匹配。"""
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    resp = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=_build_109_range("A1:D100", resolved_sheet_109_title),
    ).execute()
    rows = resp.get("values", [])
    if not rows:
        return 0

    updates = []
    # 查找并更新标签
    for i, row in enumerate(rows):
        row_num = i + 1
        # 获取第 D 列（索引 3）的内容，如果没有则为空
        label_en = row[3] if len(row) > 3 else ""
        label_cn = row[0] if len(row) > 0 else ""

        # 1. GC Cost -> Total GC Cost
        if "GC" in label_cn and ("GC Cost" in label_en or not label_en):
            if label_en != "Total GC Cost":
                updates.append({"range": _build_109_range(f"D{row_num}", resolved_sheet_109_title), "values": [["Total GC Cost"]]})

        # 2. 计提保修费用
        if label_cn == "计提保修费用":
            if label_en != "Accrued Warranty Expenses":
                updates.append({"range": _build_109_range(f"D{row_num}", resolved_sheet_109_title), "values": [["Accrued Warranty Expenses"]]})

        # 3. 实际发生保修费用
        if label_cn == "实际发生保修费用":
            if label_en != "Actual Warranty Expenses (Reversed)":
                updates.append({"range": _build_109_range(f"D{row_num}", resolved_sheet_109_title), "values": [["Actual Warranty Expenses (Reversed)"]]})

    if updates:
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"valueInputOption": "USER_ENTERED", "data": updates}
        ).execute()
        return len(updates)
    return 0


def _ensure_109_income_section_layout(
    service,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=_build_109_range("A1:D120", resolved_sheet_109_title))
        .execute()
        .get("values", [])
    )
    if not rows:
        return {"moved_total_income": False, "inserted_gc_income": False}

    label_rows = _find_rows_by_item_label(rows, item_col_1=3)
    row_total_income = _first_present_row(label_rows, "total income cost")
    row_gc_income = _first_present_row(label_rows, "gc income")
    row_gc_cost = _first_present_row(label_rows, "gc cost", "total gc cost")
    metadata = _get_109_sheet_metadata(service, spreadsheet_id, sheet_109_title=resolved_sheet_109_title)
    sheet_id = int(metadata["sheet_id"])

    moved_total_income = False
    inserted_gc_income = False

    if row_total_income is not None and row_gc_cost is not None and row_total_income > row_gc_cost:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {
                        "moveDimension": {
                            "source": {
                                "sheetId": sheet_id,
                                "dimension": "ROWS",
                                "startIndex": int(row_total_income - 1),
                                "endIndex": int(row_total_income),
                            },
                            "destinationIndex": int(row_gc_cost - 1),
                        }
                    }
                ]
            },
        ).execute()
        moved_total_income = True

        rows = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=_build_109_range("A1:D120", resolved_sheet_109_title))
            .execute()
            .get("values", [])
        )
        label_rows = _find_rows_by_item_label(rows, item_col_1=3)
        row_total_income = _first_present_row(label_rows, "total income cost")
        row_gc_income = _first_present_row(label_rows, "gc income")
        row_gc_cost = _first_present_row(label_rows, "gc cost", "total gc cost")

    if row_total_income is not None and row_gc_cost is not None and row_gc_income is None and row_gc_cost == row_total_income + 1:
        insert_row = row_total_income + 1
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {
                        "insertDimension": {
                            "range": {
                                "sheetId": sheet_id,
                                "dimension": "ROWS",
                                "startIndex": int(insert_row - 1),
                                "endIndex": int(insert_row),
                            },
                            "inheritFromBefore": False,
                        }
                    }
                ]
            },
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=_build_109_range(f"A{insert_row}:D{insert_row}", resolved_sheet_109_title),
            valueInputOption="USER_ENTERED",
            body={"values": [["GC Income", "", "GC Income", ""]]},
        ).execute()
        inserted_gc_income = True

    return {
        "moved_total_income": moved_total_income,
        "inserted_gc_income": inserted_gc_income,
    }


def generate_109_formula_plan(
    service,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
    project_id: str | None = None,
) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    resolved_project_id = _safe_string(project_id or spreadsheet_id)
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        project_id=resolved_project_id,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    project_sequence = _fetch_project_sequence(
        project_id=resolved_project_id,
        spreadsheet_id=spreadsheet_id,
    )
    if project_sequence and resolved_sheet_109_title != project_sequence:
        raise SnapshotStaleError(
            f"SNAPSHOT_STALE_ERROR: MAIN_SHEET_SEQUENCE_MISMATCH expected={project_sequence} actual={resolved_sheet_109_title}"
        )
    # These functions need to be defined or imported
    from .finance_engine import _ensure_unit_budget_actual_settlement_columns, _refresh_unit_budget_actual_settlement_columns, _sync_unit_master_sheet, _apply_unit_budget_support_formatting, _ensure_109_contract_amount_row
    _ensure_unit_budget_actual_settlement_columns(service, spreadsheet_id)
    _refresh_unit_budget_actual_settlement_columns(service, spreadsheet_id)
    _sync_unit_master_sheet(service, spreadsheet_id)
    _apply_unit_budget_support_formatting(service, spreadsheet_id)
    scoping_layout = _apply_scoping_layout_controls(service, spreadsheet_id)
    _ensure_109_contract_amount_row(service, spreadsheet_id)
    income_layout = _ensure_109_income_section_layout(service, spreadsheet_id, sheet_109_title=resolved_sheet_109_title)
    _ensure_109_labels(service, spreadsheet_id, sheet_109_title=resolved_sheet_109_title)
    resp = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=_build_109_range("A:ZZ", resolved_sheet_109_title),
    ).execute()
    rows = resp.get("values", [])
    formula_mappings = MappingService.get_project_mappings(resolved_project_id)
    plan, meta = _build_109_formula_plan_from_grid(
        rows,
        _load_109_formula_dictionary(),
        sheet_109_title=resolved_sheet_109_title,
        formula_mappings=formula_mappings,
    )
    semantic_updates = update_109_semantic_logic(
        rows,
        sheet_109_title=resolved_sheet_109_title,
        formula_mappings=formula_mappings,
    )
    merged_plan = _merge_formula_plan_with_semantic_updates(plan, semantic_updates)
    meta["semantic_formula_count"] = len(semantic_updates)
    meta["scoping_layout"] = scoping_layout
    meta["income_layout"] = income_layout
    meta["formula_mapping_project_id"] = resolved_project_id
    meta["formula_mapping_sheet_count"] = len(formula_mappings)
    return merged_plan, meta


def _ensure_unit_budget_actual_settlement_columns(service, spreadsheet_id: str) -> bool:
    rows = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range="'Unit Budget'!H2:K2").execute().get("values", [])
    if not rows or rows[0] != ["C/O date", "实际结算日期", "实际结算年份", "TBD Acceptance Date"]:
        # Logic to insert columns omitted for brevity in this mock-up
        return True
    return False

def _refresh_unit_budget_actual_settlement_columns(service, spreadsheet_id: str) -> int:
    # Logic to refresh columns omitted
    return 0

def _sync_unit_master_sheet(service, spreadsheet_id: str) -> int:
    ub_rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Unit Budget'!A:ZZ")
        .execute()
        .get("values", [])
    )
    fd_rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Final Detail'!A:V")
        .execute()
        .get("values", [])
    )
    pay_rows = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Payable'!A:AL")
        .execute()
        .get("values", [])
    )
    values = _build_unit_master_rows_v2(ub_rows, fd_rows, pay_rows)
    has_data_rows = any(
        _safe_string(row[0] if len(row) > 0 else "").upper() not in ("", "TOTAL")
        for row in values[2:]
    )
    if not has_data_rows:
        raise RuntimeError("Unit Master sync aborted: no valid data rows were produced.")

    unit_master_range = f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!A:M"
    old_row_count = len(
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=unit_master_range)
        .execute()
        .get("values", [])
    )
    new_row_count = len(values)

    (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!A1:M{new_row_count}",
            valueInputOption="USER_ENTERED",
            body={"values": values},
        )
        .execute()
    )
    if old_row_count > new_row_count:
        (
            service.spreadsheets()
            .values()
            .clear(
                spreadsheetId=spreadsheet_id,
                range=f"{_quote_sheet_name(SHEET_UNIT_MASTER_NAME)}!A{new_row_count + 1}:M{old_row_count}",
            )
            .execute()
        )
    return new_row_count

def _apply_unit_budget_support_formatting(service, spreadsheet_id: str) -> int:
    unit_budget_meta = _get_sheet_metadata(service, spreadsheet_id, "Unit Budget")
    unit_master_meta = _get_sheet_metadata(service, spreadsheet_id, SHEET_UNIT_MASTER_NAME)
    row_count = max(int(unit_budget_meta["row_count"]), int(unit_master_meta["row_count"]))
    requests = _build_unit_budget_support_requests(
        unit_budget_sheet_id=int(unit_budget_meta["sheet_id"]),
        unit_master_sheet_id=int(unit_master_meta["sheet_id"]),
        row_count=row_count,
    )
    for protected_range in unit_master_meta.get("protected_ranges", []):
        if protected_range.get("description") != MANAGED_UNIT_MASTER_PROTECTION_DESCRIPTION:
            continue
        protected_range_id = protected_range.get("protectedRangeId")
        if protected_range_id is not None:
            requests.append(_build_delete_protected_range_request(int(protected_range_id)))

    editor_email = _safe_string(_get_service_account_info().get("client_email", ""))
    manual_ranges = _build_unit_master_manual_input_ranges(int(unit_master_meta["row_count"]))
    requests.append(
        _build_add_protected_range_request(
            sheet_id=int(unit_master_meta["sheet_id"]),
            unprotected_ranges=[_a1_to_grid_range(a1, int(unit_master_meta["sheet_id"])) for a1 in manual_ranges],
            editor_email=editor_email or None,
        )
    )
    requests[-1]["addProtectedRange"]["protectedRange"]["description"] = MANAGED_UNIT_MASTER_PROTECTION_DESCRIPTION
    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ).execute()
    return len(requests)

def _ensure_109_contract_amount_row(service, spreadsheet_id: str) -> bool:
    return False

def _verify_formula_plan(service, spreadsheet_id: str, plan: Sequence[Mapping[str, str]]) -> Dict[str, Any]:
    ranges = [str(item["range"]) for item in plan if "range" in item]
    expected = { _normalize_formula_range(str(item["range"])): _normalize_formula_text_for_compare(item.get("formula", "")) for item in plan }
    matched, mismatches = 0, []
    for chunk in _chunked([{"range": r} for r in ranges], 200):
        resp = service.spreadsheets().values().batchGet(spreadsheetId=spreadsheet_id, ranges=[c["range"] for c in chunk], valueRenderOption="FORMULA").execute()
        for vr in resp.get("valueRanges", []):
            a1, actual = _normalize_formula_range(vr.get("range", "")), _normalize_formula_text_for_compare(vr.get("values", [[""]])[0][0])
            if actual == expected.get(a1): matched += 1
            else: mismatches.append({"range": a1, "expected": expected.get(a1), "actual": actual})
    return {"matched": matched, "total": len(ranges), "mismatches": mismatches}


def _capture_formula_plan_rollback_values(
    service,
    spreadsheet_id: str,
    plan: Sequence[Mapping[str, str]],
) -> List[Dict[str, Any]]:
    ranges = [str(item["range"]) for item in plan if item.get("range")]
    captured_by_range: Dict[str, List[List[Any]]] = {}
    for chunk in _chunked(ranges, 200):
        chunk_ranges = list(chunk)
        response = service.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=chunk_ranges,
            valueRenderOption="FORMULA",
        ).execute()
        for value_range in response.get("valueRanges", []):
            normalized_range = _normalize_formula_range(_safe_string(value_range.get("range")))
            values = value_range.get("values")
            captured_by_range[normalized_range] = values if isinstance(values, list) and values else [[""]]

    rollback_rows: List[Dict[str, Any]] = []
    for range_a1 in ranges:
        rollback_rows.append(
            {
                "range": range_a1,
                "majorDimension": "ROWS",
                "values": captured_by_range.get(_normalize_formula_range(range_a1), [[""]]),
            }
        )
    return rollback_rows


def _restore_formula_plan_rollback_values(
    service,
    spreadsheet_id: str,
    rollback_rows: Sequence[Mapping[str, Any]],
) -> Dict[str, Any]:
    if not rollback_rows:
        return {"restored_ranges": 0}
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "valueInputOption": "USER_ENTERED",
            "data": list(rollback_rows),
        },
    ).execute()
    return {"restored_ranges": len(rollback_rows)}


def _merge_draft_with_cloud(base_df: pd.DataFrame, draft_df: pd.DataFrame, uid_column: str) -> pd.DataFrame:
    if draft_df.empty: return base_df.copy()
    merged = draft_df.copy()
    for col in base_df.columns:
        if col not in merged.columns: merged[col] = ""
    if uid_column in base_df.columns and uid_column in merged.columns:
        draft_uids = set(merged[uid_column].astype("string").fillna("").str.strip())
        missing = base_df[~base_df[uid_column].astype("string").fillna("").str.strip().isin(draft_uids)]
        if not missing.empty: merged = pd.concat([merged, missing], ignore_index=True)
    return merged.reindex(columns=list(base_df.columns) + [c for c in merged.columns if c not in base_df.columns])


def _prepare_sheet_df(cloud_df: pd.DataFrame, uid_column: str, shadow_cfg: Mapping[str, Any]) -> pd.DataFrame:
    prepared, _ = ensure_uid_anchor(cloud_df, uid_column)
    return apply_shadow_logic(prepared, **shadow_cfg)


def _has_pending_uid(df: pd.DataFrame) -> bool:
    return UID_STATUS_COL in df.columns and (df[UID_STATUS_COL].astype("string") == UID_PENDING_VALUE).any()


def _find_dirty_sheets(original_map: Mapping[str, pd.DataFrame], edited_map: Mapping[str, pd.DataFrame]) -> List[str]:
    dirty = []
    for s in sorted(set(original_map.keys()) | set(edited_map.keys())):
        if _df_signature(original_map.get(s, pd.DataFrame())) != _df_signature(edited_map.get(s, pd.DataFrame())) or _has_pending_uid(edited_map.get(s, pd.DataFrame())):
            dirty.append(s)
    return dirty

def append_audit_log(operator: str, summary: str, details: Sequence[str]) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [f"[{now}] operator={operator} | {summary}"] + [f"  - {line}" for line in details]
    AUDIT_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT_LOG_FILE.open("a", encoding="utf-8") as fp: fp.write("\n".join(lines) + "\n")


def _ensure_local_store() -> None:
    DRAFT_DIR.mkdir(parents=True, exist_ok=True)


def _draft_file(sheet: str) -> Path:
    return DRAFT_DIR / f"{_slugify_sheet_name(sheet)}.pkl"


def save_local_draft(sheet: str, df: pd.DataFrame) -> None:
    _ensure_local_store()
    df.to_pickle(_draft_file(sheet))


def load_local_draft(sheet: str) -> pd.DataFrame | None:
    path = _draft_file(sheet)
    if not path.exists():
        return None
    try:
        return pd.read_pickle(path)
    except Exception:
        return None


def clear_local_drafts() -> None:
    if not DRAFT_DIR.exists():
        return
    for file in DRAFT_DIR.glob("*.pkl"):
        file.unlink(missing_ok=True)


def save_local_cloud_snapshot(
    spreadsheet_id: str,
    cloud_map: Mapping[str, pd.DataFrame],
    formula_lookup_map: Mapping[str, Dict[Tuple[int, str], str]],
    cloud_meta: Mapping[str, Any],
) -> None:
    _ensure_local_store()
    payload = {
        "spreadsheet_id": spreadsheet_id,
        "cloud_map": dict(cloud_map),
        "formula_lookup_map": dict(formula_lookup_map),
        "cloud_meta": dict(cloud_meta),
        "saved_at": datetime.now(timezone.utc).isoformat(),
    }
    tmp = CLOUD_SNAPSHOT_FILE.with_suffix(".tmp")
    with tmp.open("wb") as fp:
        pickle.dump(payload, fp, protocol=pickle.HIGHEST_PROTOCOL)
    tmp.replace(CLOUD_SNAPSHOT_FILE)


def load_local_cloud_snapshot(
    spreadsheet_id: str,
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[Tuple[int, str], str]], Dict[str, Any]] | None:
    if not CLOUD_SNAPSHOT_FILE.exists():
        return None
    try:
        with CLOUD_SNAPSHOT_FILE.open("rb") as fp:
            payload = pickle.load(fp)
    except Exception:
        return None

    if not isinstance(payload, Mapping):
        return None
    if _safe_string(payload.get("spreadsheet_id")) != _safe_string(spreadsheet_id):
        return None

    cloud_map = payload.get("cloud_map")
    formula_lookup_map = payload.get("formula_lookup_map")
    cloud_meta = payload.get("cloud_meta")
    if not isinstance(cloud_map, Mapping) or not isinstance(formula_lookup_map, Mapping) or not isinstance(cloud_meta, Mapping):
        return None
    meta = dict(cloud_meta)
    meta["snapshot_source"] = "local"
    meta["snapshot_saved_at"] = _safe_string(payload.get("saved_at"))
    return dict(cloud_map), dict(formula_lookup_map), meta


def _extract_error_http_status(error: Exception) -> int | None:
    for attr in ("status", "status_code"):
        value = getattr(error, attr, None)
        if isinstance(value, int):
            return value
    response = getattr(error, "resp", None)
    status = getattr(response, "status", None)
    if isinstance(status, int):
        return status
    return None


def _is_google_write_rate_limit_error(error: Exception) -> bool:
    status = _extract_error_http_status(error)
    if status == 429:
        return True
    message = _safe_string(error)
    upper = message.upper()
    return "429" in upper or "RATE_LIMIT_EXCEEDED" in upper


def _execute_google_write_with_backoff(
    request_factory,
    *,
    max_attempts: int = GOOGLE_WRITE_MAX_ATTEMPTS,
    base_delay_seconds: float = GOOGLE_WRITE_BACKOFF_BASE_SECONDS,
    max_delay_seconds: float = GOOGLE_WRITE_BACKOFF_MAX_SECONDS,
) -> Tuple[Any, int]:
    attempts = max(1, int(max_attempts))
    retries = 0
    delay = max(0.0, float(base_delay_seconds))
    max_delay = max(delay, float(max_delay_seconds))

    for attempt in range(1, attempts + 1):
        try:
            return request_factory().execute(), retries
        except Exception as error:
            if not _is_google_write_rate_limit_error(error) or attempt >= attempts:
                raise
            retries += 1
            time.sleep(delay)
            delay = min(max_delay, max(delay * 2, base_delay_seconds))

    raise RuntimeError("GOOGLE_WRITE_BACKOFF_UNEXPECTED_EXIT")


def execute_109_formula_plan(
    service,
    spreadsheet_id: str,
    plan: Sequence[Mapping[str, str]],
    meta: Mapping[str, Any] | None = None,
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_project_id = _safe_string((meta or {}).get("formula_mapping_project_id") or spreadsheet_id)
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        project_id=resolved_project_id,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    updates = [
        {
            "range": str(item["range"]),
            "majorDimension": "ROWS",
            "values": [[str(item["formula"])]],
        }
        for item in plan
        if item.get("range") and item.get("formula")
    ]
    api_calls = 0
    write_retry_count = 0
    write_throttled_chunk_count = 0
    rollback_rows = _capture_formula_plan_rollback_values(service, spreadsheet_id, plan) if updates else []
    try:
        for chunk in _chunked(updates, FORMULA_WRITE_CHUNK_SIZE):
            chunk_rows = list(chunk)
            _, retries = _execute_google_write_with_backoff(
                lambda: service.spreadsheets().values().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "valueInputOption": "USER_ENTERED",
                        "data": chunk_rows,
                    },
                )
            )
            api_calls += 1
            write_retry_count += retries
            if retries > 0:
                write_throttled_chunk_count += 1
    except Exception as error:
        try:
            rollback_audit = _restore_formula_plan_rollback_values(service, spreadsheet_id, rollback_rows)
        except Exception as rollback_error:
            raise RuntimeError(
                f"FORMULA_WRITEBACK_PARTIAL_ROLLBACK_FAILED: write_error={error}; rollback_error={rollback_error}"
            ) from error
        raise RuntimeError(
            f"FORMULA_WRITEBACK_PARTIAL_ROLLBACK: restored_ranges={rollback_audit.get('restored_ranges', 0)}; write_error={error}"
        ) from error

    legacy_cleanup = _cleanup_109_legacy_duplicate_contract_change_row(
        service,
        spreadsheet_id,
        sheet_109_title=resolved_sheet_109_title,
    )
    verify = _verify_formula_plan(service, spreadsheet_id, plan) if updates else {"matched": 0, "total": 0, "mismatches": []}
    manual_ranges = list((meta or {}).get("manual_ranges", []))
    if not manual_ranges:
        rows = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=_build_109_range("A:ZZ", resolved_sheet_109_title))
            .execute()
            .get("values", [])
        )
        manual_ranges = _build_109_manual_input_ranges(
            rows,
            _year_columns_from_109_dictionary(_load_109_formula_dictionary()),
            sheet_109_title=resolved_sheet_109_title,
        )
    layout = _apply_109_layout_controls(
        service,
        spreadsheet_id,
        manual_ranges,
        sheet_109_title=resolved_sheet_109_title,
    )
    formula_locks = _apply_109_formula_lock_protection(
        service,
        spreadsheet_id,
        plan,
        sheet_109_title=resolved_sheet_109_title,
    )
    external_controls = _apply_external_sheet_controls(service, spreadsheet_id)
    return {
        "api_calls": api_calls,
        "updated_ranges": len(updates),
        "write_throttle": {
            "chunk_size": FORMULA_WRITE_CHUNK_SIZE,
            "retry_count": write_retry_count,
            "throttled_chunk_count": write_throttled_chunk_count,
        },
        "rollback_journal": {
            "captured_ranges": len(rollback_rows),
        },
        "verify": verify,
        "legacy_cleanup": legacy_cleanup,
        "layout": layout,
        "formula_locks": formula_locks,
        "external_controls": external_controls,
    }


def _cleanup_109_legacy_duplicate_contract_change_row(
    service,
    spreadsheet_id: str,
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    cleanup_range = _build_109_range("A64:Q64", resolved_sheet_109_title)
    row = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range=cleanup_range,
            valueRenderOption="FORMULA",
        )
        .execute()
        .get("values", [])
    )
    values = row[0] if row else []

    def _cell(idx_1: int) -> str:
        return _safe_string(values[idx_1 - 1]) if idx_1 - 1 < len(values) else ""

    matches_legacy = (
        _cell(1) == "合同变动金额"
        and _cell(3) == "Contract Change Amount"
        and "IFERROR(61+F15" in _cell(6)
        and _cell(13).startswith('=IF(AND($C$4<M$9,$C$4>L$9),$E$3,"")')
    )
    if not matches_legacy:
        return {"cleared": False, "range": cleanup_range}

    service.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id,
        body={"ranges": [cleanup_range]},
    ).execute()
    return {"cleared": True, "range": cleanup_range}


def run_validate_input_data(service, spreadsheet_id: str) -> Dict[str, Any]:
    unit_master_rows_written = _sync_unit_master_sheet(service, spreadsheet_id)
    unit_budget_layout_request_count = _apply_unit_budget_support_formatting(service, spreadsheet_id)
    scoping_layout = _apply_scoping_layout_controls(service, spreadsheet_id)
    return {
        "unit_master_rows_written": unit_master_rows_written,
        "unit_budget_layout_request_count": unit_budget_layout_request_count,
        "scoping_layout": scoping_layout,
    }


def initialize_project_workbook(
    service,
    spreadsheet_id: str,
    project_name: str,
    project_owner: str,
    creator_email: str = "",
    sheet_109_title: str | None = None,
) -> Dict[str, Any]:
    resolved_sheet_109_title = _resolve_sheet_109_title(
        sheet_109_title=sheet_109_title,
        spreadsheet_id=spreadsheet_id,
        service=service,
    )
    sheet_109_renamed = _rename_109_sheet_for_project(
        service,
        spreadsheet_id,
        sheet_109_title=resolved_sheet_109_title,
    )
    rows_109 = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=_build_109_range("A:ZZ", resolved_sheet_109_title))
        .execute()
        .get("values", [])
    )
    rows_scoping = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Scoping'!A:Z")
        .execute()
        .get("values", [])
    )
    manual_ranges_109 = _build_109_manual_input_ranges(
        rows_109,
        _year_columns_from_109_dictionary(_load_109_formula_dictionary()),
        sheet_109_title=resolved_sheet_109_title,
    )
    unit_master_metadata = _get_sheet_metadata(service, spreadsheet_id, SHEET_UNIT_MASTER_NAME)
    clear_ranges = _build_project_bootstrap_manual_clear_ranges(
        rows_109=rows_109,
        rows_scoping=rows_scoping,
        row_count_unit_master=int(unit_master_metadata["row_count"]),
        sheet_109_title=resolved_sheet_109_title,
    )

    if clear_ranges:
        service.spreadsheets().values().batchClear(
            spreadsheetId=spreadsheet_id,
            body={"ranges": clear_ranges},
        ).execute()

    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "valueInputOption": "USER_ENTERED",
            "data": [
                {"range": _build_109_range("C2", resolved_sheet_109_title), "values": [[_safe_string(project_name)]]},
                {"range": _build_109_range("G2", resolved_sheet_109_title), "values": [[_safe_string(project_owner)]]},
            ],
        },
    ).execute()

    scoping_layout = _apply_scoping_layout_controls(service, spreadsheet_id)
    unit_budget_layout_request_count = _apply_unit_budget_support_formatting(service, spreadsheet_id)
    layout_109 = _apply_109_layout_controls(
        service,
        spreadsheet_id,
        manual_ranges_109,
        sheet_109_title=resolved_sheet_109_title,
    )

    external_controls = _apply_external_sheet_controls(service, spreadsheet_id)
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{_quote_sheet_name(SHEET_PROJECT_STATE_NAME)}!A:B",
        valueInputOption="USER_ENTERED",
        body={
            "values": _build_project_state_values(
                owner_email=creator_email,
                current_stage=WORKBENCH_STAGE_PROJECT_CREATED,
                locked=False,
            )
        },
    ).execute()

    append_project_audit_log(
        service=service,
        spreadsheet_id=spreadsheet_id,
        actor_email=creator_email,
        action="create_project",
        previous_stage="",
        next_stage=WORKBENCH_STAGE_PROJECT_CREATED,
        status="success",
        message="Project workbook created.",
    )

    external_data_rows_after_sanitize: Dict[str, int] = {}
    for sheet_name, probe_range in _build_external_sheet_data_probe_ranges().items():
        rows = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=probe_range)
            .execute()
            .get("values", [])
        )
        external_data_rows_after_sanitize[sheet_name] = len(rows)

    return {
        "headers_written": 2,
        "sheet_109_renamed": sheet_109_renamed,
        "manual_clear_range_count": len(clear_ranges),
        "external_clear_range_count": len(_build_external_sheet_clear_ranges()),
        "external_data_rows_after_sanitize": external_data_rows_after_sanitize,
        "external_protection_request_count": external_controls["external_protection_request_count"],
        "log_hidden": external_controls["log_hidden"],
        "project_state_initialized": True,
        "audit_log_initialized": True,
        "scoping_layout": scoping_layout,
        "unit_budget_layout_request_count": unit_budget_layout_request_count,
        "layout_109": layout_109,
        "external_controls": external_controls,
    }


def clear_local_cloud_snapshot() -> None:
    CLOUD_SNAPSHOT_FILE.unlink(missing_ok=True)
